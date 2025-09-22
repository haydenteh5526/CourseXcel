import os, logging, pytz
from app.models import Rate
from collections import defaultdict
from copy import copy
from datetime import datetime
from flask import current_app
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO)

# Convert a date object to DD/MM/YYYY format
def format_date(date_obj):
    try:
        if isinstance(date_obj, str):
            date_obj = datetime.strptime(date_obj, "%Y-%m-%d").date()
        return date_obj.strftime('%d/%m/%Y') if date_obj else ''
    except Exception as e:
        logging.error(f"Date formatting error: {e}")
        return ''

# Convert a date object to DD MMM YYYY format
def format_file_date(date_obj):
    try:
        if isinstance(date_obj, str):
            date_obj = datetime.strptime(date_obj, "%Y-%m-%d").date()
        # Example: 1 Sep 2025
        return date_obj.strftime('%-d %b %Y') if date_obj else ''
    except Exception as e:
        logging.error(f"File Date formatting error: {e}")
        return ''

def get_local_date_str(timezone_str='Asia/Kuala_Lumpur'):
    tz = pytz.timezone(timezone_str)
    now = datetime.now(tz)
    return now.strftime('%d/%m/%Y')

# =============== Requisition Excel =============== #
def generate_requisition_excel(department_code, name, designation, ic_number, subject_level, course_details, po_name, head_name, dean_name, ad_name, hr_name):
    try:
        # Load template and define paths
        template_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 
                                   "files", 
                                   "Part-Time Lecturer Requisition Form - template.xlsx")
        output_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), "temp")
        output_filename = f"Part-Time Lecturer Requisition Form - {name} ({subject_level}).xlsx"
        output_path = os.path.join(output_folder, output_filename)

        # Ensure the output directory exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Load the workbook and active sheet
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active

        # Insert lecturer and department details
        template_ws['C5'].value = department_code
        template_ws['C6'].value = name
        template_ws['C7'].value = designation
        template_ws['H6'].value = ic_number

        # Store the template of the first course record (A9:L22)
        first_record_template = []
        for row in range(9, 23):
            row_data = []
            for col in range(1, 13):  # Columns A to L
                cell = template_ws.cell(row=row, column=col)
                row_data.append({
                    'value': cell.value,
                    'style': copy(cell._style),
                    'formula': cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
                })
            first_record_template.append(row_data)

        # Track total cost cells to later calculate grand total
        total_cost_cells = ['I20']  # First record's total cost cell

        # Process each course in course_details
        for index, course in enumerate(course_details):
            if index == 0:
                # Use existing template space for the first course
                insert_requisition_record(template_ws, course, 9)
            else:
                # Determine row to insert new course
                insert_point = 23 + (14 * (index - 1))
                
                # Insert 14 new rows for the new course section
                template_ws.insert_rows(insert_point, 14)
                
                # Copy layout and styles from the first template block
                copy_requisition_structure(template_ws, first_record_template, insert_point)
                
                # Fill in course data
                insert_requisition_record(template_ws, course, insert_point)
                
                # Update formulas for cost in new section
                update_requisition_formulas(template_ws, insert_point)
                
                # Add this course's cost cell to the list
                total_cost_cells.append(f'I{insert_point + 11}')

        # Write total cost formula summing all tracked cells
        final_total_row = 23 + (14 * (len(course_details) - 1))
        template_ws[f'I{final_total_row}'].value = f'=SUM({",".join(total_cost_cells)})'

        # Maps for determining where to insert final signature block
        row_map = {1: 29, 2: 43, 3: 57, 4: 71}
        merge_row_map = {1: 27, 2: 41, 3: 55, 4: 69}
        num_courses = len(course_details)
        start_row = row_map.get(num_courses)
        merge_row = merge_row_map.get(num_courses)

        # Insert and style signature blocks if applicable
        if start_row and merge_row:
            # Merge and center cells for signature layout
            template_ws.merge_cells(f'B{merge_row}:B{merge_row + 1}')
            template_ws.merge_cells(f'E{merge_row}:E{merge_row + 1}')
            template_ws.merge_cells(f'G{merge_row}:G{merge_row + 1}')
            template_ws.merge_cells(f'I{merge_row}:I{merge_row + 1}')
            template_ws.merge_cells(f'K{merge_row}:K{merge_row + 1}')

            # Center align the merged cells
            for col in ['B', 'E', 'G', 'I', 'K']:
                cell = template_ws[f'{col}{merge_row}']
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Populate signature block with names and date
            template_ws[f'B{start_row}'].value = f"Name: {po_name}"
            template_ws[f'B{start_row + 1}'].value = f"Date: {get_local_date_str()}"
            template_ws[f'E{start_row}'].value = f"Name: {head_name}"
            template_ws[f'G{start_row}'].value = f"Name: {dean_name}"
            template_ws[f'I{start_row}'].value = f"Name: {ad_name}"
            template_ws[f'K{start_row}'].value = f"Name: {hr_name}"

        # Protect the worksheet to make it read-only
        template_ws.protection.sheet = True
        
        # Save the final output Excel file
        template_wb.save(output_path)
        return output_path, merge_row

    except Exception as e:
        logging.error(f"Error generating Excel file: {e}")
        raise

# Copy the record structure from the requisition template to a new location in the worksheet
def copy_requisition_structure(ws, template_data, start_row):
    try:
        # Loop through each row in the stored template data
        for row_idx, row_data in enumerate(template_data):
            target_row = start_row + row_idx   # Determine target row number
            
            # Loop through each column (cell) in the current row
            for col_idx, cell_data in enumerate(row_data, start=1):
                target_cell = ws.cell(row=target_row, column=col_idx)
                
                # Copy value only if it's not a formula (formulas will be set separately)
                if cell_data['value'] and not cell_data['formula']:
                    target_cell.value = cell_data['value']
                
                # Apply the original cell style
                target_cell._style = copy(cell_data['style'])
                
        logging.info(f"Successfully copied record structure to row {start_row}")
        
    except Exception as e:
        logging.error(f"Error copying record structure: {e}")
        raise

# Insert course details into the Excel worksheet
def insert_requisition_record(ws, course, start_row):
    try:
        # Calculate row positions relative to start_row
        subject_title_row = start_row + 1  # Row 10 for first record
        subject_level_row = start_row + 2  # Row 11 for first record
        teaching_period_row = start_row + 4  # Row 13 for first record
        
        # Category rows (Lecture, Tutorial, Practical, Blended)
        category_start = start_row + 6  # Row 15 for first record
        
        # Insert basic course information
        ws[f'C{subject_title_row}'].value = course['subject_title']
        ws[f'I{subject_title_row}'].value = course['subject_code']
        ws[f'C{subject_level_row}'].value = course['subject_level']
        
        # Insert teaching period dates
        start_date = format_date(course['start_date'])
        end_date = format_date(course['end_date'])
        ws[f'C{teaching_period_row}'].value = f"From {start_date} to {end_date}"
        
        # Insert teaching hours
        ws[f'D{category_start}'].value = course['lecture_hours']
        ws[f'D{category_start + 1}'].value = course['tutorial_hours']
        ws[f'D{category_start + 2}'].value = course['blended_hours']
        ws[f'D{category_start + 3}'].value = course['practical_hours']
        
        # Insert teaching weeks
        ws[f'G{category_start}'].value = course['lecture_weeks']
        ws[f'G{category_start + 1}'].value = course['tutorial_weeks']
        ws[f'G{category_start + 2}'].value = course['blended_weeks']
        ws[f'G{category_start + 3}'].value = course['practical_weeks']
        
        # Insert hourly rate
        ws[f'D{start_row + 11}'].value = course['hourly_rate']  # Row 20 for first record
        
        logging.info(f"Successfully inserted course data starting at row {start_row}")
        
    except Exception as e:
        logging.error(f"Error inserting record: {e}")
        raise

# Update Excel formulas for a record block
def update_requisition_formulas(ws, start_row):
    try:
        # Calculate formula positions
        category_start = start_row + 6  # Row 15 for first record (categories start)
        total_row = start_row + 11      # Row 20 for first record (totals row)
        
        # Update category total formulas (hours × weeks)
        # Example: For first record, updates I15 to I18
        for i in range(4):  # 4 categories: Lecture, Tutorial, Practical, Blended
            row = category_start + i
            # Formula: hours × weeks (e.g., =D15*G15)
            ws[f'I{row}'].value = f'=D{row}*G{row}'
        
        # Update total hours formula
        # Example: For first record, G20 = SUM(I15:I18)
        ws[f'G{total_row}'].value = f'=SUM(I{category_start}:I{category_start+3})'
        
        # Update total cost formula
        # Example: For first record, J20 = D20*G20
        ws[f'I{total_row}'].value = f'=D{total_row}*G{total_row}'
        
        logging.info(f"Successfully updated formulas starting at row {start_row}")
        
    except Exception as e:
        logging.error(f"Error updating record formulas: {e}")
        raise

# =============== Claim Excel =============== #
def generate_claim_excel(name, department_code, subject_level, claim_details, po_name, head_name, dean_name, hr_name):
    try:
        # Load template
        template_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 
                                   "files", 
                                   "Part-Time Lecturer Claim Form - template.xlsx")
        output_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), "temp")
        output_filename = f"Part-Time Lecturer Claim Form - {name} ({subject_level}).xlsx"
        output_path = os.path.join(output_folder, output_filename)

        # Ensure output directory exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Load workbook
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active

        # Add image to worksheet
        template_ws.merge_cells('G3:G5')
        logo_path = os.path.join(current_app.root_path, 'static', 'img', 'Form INTI Logo.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 160
            img.height = 100
            img.anchor = 'G3'
            template_ws.add_image(img)

        # Default
        first_rate_amount = 0

        if claim_details:
            first_claim = claim_details[0]
            rate_id = first_claim.get('rate_id')
            if rate_id:
                r = Rate.query.get(rate_id)
                first_rate_amount = r.amount if r else 0

        # Insert lecturer details
        template_ws['B5'].value = name
        template_ws['B6'].value = department_code
        template_ws['B7'].value = subject_level
        template_ws['B8'].value = first_rate_amount

        # Insert claim details starting from row 14
        start_row = 12
        max_rows = 40

        for index, claim in enumerate(claim_details[:max_rows]):
            row = start_row + index
            template_ws[f"A{row}"].value = format_date(claim['date'])
            template_ws[f"B{row}"].value = claim['subject_code']
            template_ws[f"C{row}"].value = claim['lecture_hours']
            template_ws[f"D{row}"].value = claim['tutorial_hours']
            template_ws[f"E{row}"].value = claim['practical_hours']
            template_ws[f"F{row}"].value = claim['blended_hours']
            template_ws[f"G{row}"].value = claim['remarks']

        # Insert total claimed hours formulas
        template_ws["C52"] = "=SUM(C12:C51)"
        template_ws["D52"] = "=SUM(D12:D51)"
        template_ws["E52"] = "=SUM(E12:E51)"
        template_ws["F52"] = "=SUM(F12:F51)"

        # Insert rounded total payment formulas based on hourly rate
        template_ws["B57"] = "=ROUND(B8*C52, 2)"
        template_ws["B58"] = "=ROUND(B8*D52, 2)"
        template_ws["B59"] = "=ROUND(B8*E52, 2)"
        template_ws["B60"] = "=ROUND(B8*F52, 2)"

        # Insert grand total payment (sum of all above payments)
        template_ws["B61"] = "=SUM(B57:B60)"

        # Handle signature cells
        sign_col = 65
        name_col = 67

        # Merge the rows
        template_ws.merge_cells(f'A{sign_col}:A{sign_col + 1}')
        template_ws.merge_cells(f'B{sign_col}:B{sign_col + 1}')
        template_ws.merge_cells(f'C{sign_col}:C{sign_col + 1}')
        template_ws.merge_cells(f'E{sign_col}:E{sign_col + 1}')
        template_ws.merge_cells(f'G{sign_col}:G{sign_col + 1}')

        # Center align the merged cells
        for col in ['A', 'B', 'C', 'E' 'G']:
            cell = template_ws[f'{col}{sign_col}']  # Get the first cell of the merged range
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Fill values in correct cells
        template_ws[f'A{name_col}'].value = f"Name: {name}"
        template_ws[f'A{name_col + 1}'].value = f"Date: {get_local_date_str()}"
        template_ws[f'B{name_col}'].value = f"Name: {po_name}"
        template_ws[f'C{name_col}'].value = f"Name: {head_name}"
        template_ws[f'E{name_col}'].value = f"Name: {dean_name}"
        template_ws[f'G{name_col}'].value = f"Name: {hr_name}"

        # Protect the worksheet and make it read-only
        template_ws.protection.sheet = True
        
        # Save the file
        template_wb.save(output_path)
        return output_path, sign_col

    except Exception as e:
        logging.error(f"Error generating Excel file: {e}")
        raise

# =============== Report Excel =============== #
# Overall layout constants
OVERALL_DETAILS_START_ROW   = 7   # first detail row on overall sheet (row 6 is header)
OVERALL_SUMMARY_DATA_START  = 11  # first row of summary data

# Department layout constants
DEPT_DETAILS_START_ROW = 8        # first detail row on dept sheets (row 7 is header)
DEPT_SUMMARY_DATA_START = 12      # first row of summary data
DEPT_SUMMARY_TOTAL_COL = "B"      # default column for "Total" values in depts summary

DEPARTMENT_SHEETS = ["CADP", "CAE", "CEPS", "LCMPU", "SOBIZ", "SOC", "SOE", "SOHOS"]

COLORS = ["B7950B", "D35400", "C0392B", "27AE60", "16A085", "2980B9", "8E44AD", "7F8C8D"]

def copy_row_style(ws, src_row, dst_row):
    """
    Clone cell styles from src_row to dst_row (number formats, borders, fills, fonts, alignments).
    """
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        s = ws.cell(src_row, col)
        d = ws.cell(dst_row, col)
        if s.has_style:
            d._style = copy(s._style)

# Insert one detail record
def insert_overall_details(ws, report, start_row):
    """
    Write one detail row.
    report: dict with keys:
        department_code, lecturer_name, total_subjects,
        total_lecture_hours, total_tutorial_hours,
        total_practical_hours, total_blended_hours, rate, total_cost
    """
    ws[f'A{start_row}'].value = report.get('department_code', '')
    ws[f'B{start_row}'].value = report.get('lecturer_name', '')
    ws[f'C{start_row}'].value = report.get('total_subjects', 0)
    ws[f'D{start_row}'].value = report.get('total_lecture_hours', 0)
    ws[f'E{start_row}'].value = report.get('total_tutorial_hours', 0)
    ws[f'F{start_row}'].value = report.get('total_practical_hours', 0)
    ws[f'G{start_row}'].value = report.get('total_blended_hours', 0)
    ws[f'H{start_row}'].value = report.get('rate', 0)
    ws[f'I{start_row}'].value = report.get('total_cost', 0)

def merge_same_department(ws, first_row: int, total_rows: int, col: str = "A"):
    """
    Merge consecutive cells in `col` (e.g., 'A') when they have the same non-empty value,
    and center the text. Clears tail cells BEFORE merging to avoid MergedCell write errors.
    """
    if total_rows <= 0:
        return

    start = first_row
    end   = first_row + total_rows - 1
    r = start

    while r <= end:
        val = ws[f"{col}{r}"].value
        if val is None or str(val).strip() == "":
            r += 1
            continue

        run_start = r
        # extend while next row has the same value
        while r + 1 <= end and ws[f"{col}{r+1}"].value == val:
            r += 1
        run_end = r

        if run_end > run_start:
            # 1) clear tail cells BEFORE we merge
            for rr in range(run_start + 1, run_end + 1):
                ws[f"{col}{rr}"].value = None

            # 2) merge and center
            ws.merge_cells(f"{col}{run_start}:{col}{run_end}")
            top = ws[f"{col}{run_start}"]
            top.alignment = Alignment(horizontal="center", vertical="center")

        r += 1

# Build the Summary Table
def write_overall_summary(ws, report_details, data_start_row, put_chart=True):
    """
    Fills the Summary Table below the 'Summary Table' header.
    Columns:
      A = Department
      B = No. of Lecturers (distinct)
      C = No. of Subjects (sum)
      D = Total Cost (sum)
    """

    # Aggregate from details
    agg = defaultdict(lambda: {"lecturers": set(), "subjects": 0, "cost": 0})
    for r in report_details:
        dep = (r.get("department_code") or "-").strip()
        agg[dep]["lecturers"].add(r.get("lecturer_name", "") or "")
        agg[dep]["subjects"] += int(r.get("total_subjects") or 0)
        agg[dep]["cost"]     += int(r.get("total_cost") or 0)

    # Read existing department labels in the template (col A)
    existing_rows = []
    row = data_start_row
    while True:
        v = ws[f"A{row}"].value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            break
        existing_rows.append((row, str(v).strip()))
        row += 1

    # Fill rows for existing template departments (with 0 if missing)
    last = data_start_row - 1
    for rix, dep in existing_rows:
        stats = agg.get(dep, {"lecturers": set(), "subjects": 0, "cost": 0})
        ws[f"B{rix}"].value = len(stats["lecturers"])
        ws[f"C{rix}"].value = stats["subjects"]
        ws[f"D{rix}"].value = stats["cost"]
        last = rix

    # Append any departments from data not in template
    template_deps = {dep for _, dep in existing_rows}
    for dep, stats in sorted(agg.items()):
        if dep not in template_deps:
            last += 1
            ws.insert_rows(last, amount=1)
            ws[f"A{last}"].value = dep
            ws[f"B{last}"].value = len(stats["lecturers"])
            ws[f"C{last}"].value = stats["subjects"]
            ws[f"D{last}"].value = stats["cost"]

    first_row = data_start_row
    last_row  = last

    # Chart
    if put_chart and last_row >= first_row:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Total Cost by Department"
        chart.y_axis.title = "RM"
        chart.x_axis.title = "Department"

        # Categories (departments)
        cats = Reference(ws, min_col=1, min_row=first_row, max_row=last_row)
        # Data (Total Cost column D)
        data = Reference(ws, min_col=4, min_row=first_row, max_row=last_row)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.legend = None

        # Custom colors for each bar
        series = chart.series[0]
        
        for idx, color in enumerate(COLORS, start=0):
            dp = DataPoint(idx=idx)
            dp.graphicalProperties.solidFill = color
            series.dPt.append(dp)

        # Make chart larger
        chart.width = 20   # default ~15
        chart.height = 12  # default ~7.5

        # Add chart to sheet
        ws.add_chart(chart, "K6")

# helpers for department sheets
def group_details_by_department(report_details):
    """ { 'CADP': [row,row,...], 'CAE': [...], ... } """
    buckets = defaultdict(list)
    for r in report_details:
        dep = (r.get("department_code") or "").strip()
        buckets[dep].append(r)
    return buckets

def find_row_by_label(ws, label, col=1, start_row=1, end_row=200):
    tgt = str(label).strip().lower()
    for r in range(start_row, end_row+1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip().lower() == tgt:
            return r
    return None

def find_total_col(ws, header_row, header_text="Total"):
    tgt = str(header_text).strip().lower()
    for c in range(1, ws.max_column+1):
        if str(ws.cell(header_row, c).value or "").strip().lower() == tgt:
            return get_column_letter(c)
    return None  # caller can default

def write_department_details(ws, dept_rows):
    """
    Fill a department sheet's Details block.
    - Insert (n-1) rows starting at DEPT_DETAILS_START_ROW
    - Copy styles from the template row (DEPT_DETAILS_START_ROW)
    - Shift summary rows down if rows are inserted
    """

    n = len(dept_rows)
    if n == 0:
        return

    if n > 1:
        ws.insert_rows(DEPT_DETAILS_START_ROW + 1, amount=n - 1)

    for i, rec in enumerate(dept_rows):
        r = DEPT_DETAILS_START_ROW + i
        if i > 0:
            copy_row_style(ws, DEPT_DETAILS_START_ROW, r)
        ws[f'A{r}'].value = rec.get('lecturer_name', '')
        ws[f'B{r}'].value = rec.get('total_subjects', 0)
        ws[f'C{r}'].value = rec.get('total_lecture_hours', 0)
        ws[f'D{r}'].value = rec.get('total_tutorial_hours', 0)
        ws[f'E{r}'].value = rec.get('total_practical_hours', 0)
        ws[f'F{r}'].value = rec.get('total_blended_hours', 0)
        ws[f'G{r}'].value = rec.get('rate', 0)
        ws[f'H{r}'].value = rec.get('total_cost', 0)

def apply_colors(series, num_points, colors=COLORS):
    """Assign solid fills to the first `num_points` points."""
    for idx in range(num_points):
        dp = DataPoint(idx=idx)
        dp.graphicalProperties.solidFill = colors[idx % len(colors)]
        series.dPt.append(dp)

def write_department_summary(ws, dept_rows, total_col=None):
    """
    Robust to any number of inserted detail rows.
    Finds:
      - first_label_row = row where column A == "No. of Lecturers"
      - header_row = first_label_row - 1
      - total_col = column with header "Total" on header_row
    """
    # locate summary block for THIS sheet
    first_label_row = find_row_by_label(ws, "No. of Lecturers", col=1, start_row=8, end_row=100)
    if first_label_row is None:
        first_label_row = DEPT_SUMMARY_DATA_START  # fallback to template constant
    header_row = first_label_row - 1

    if total_col is None:
        total_col = find_total_col(ws, header_row) or DEPT_SUMMARY_TOTAL_COL

    # compute totals
    lecturers = {r.get("lecturer_name","") for r in dept_rows}
    num_lecturers = len(lecturers)
    num_subjects = sum(int(r.get("total_subjects") or 0) for r in dept_rows)
    lec_hours = sum(int(r.get("total_lecture_hours") or 0) for r in dept_rows)
    tut_hours = sum(int(r.get("total_tutorial_hours") or 0) for r in dept_rows)
    prac_hours = sum(int(r.get("total_practical_hours") or 0) for r in dept_rows)
    blend_hours = sum(int(r.get("total_blended_hours") or 0) for r in dept_rows)
    total_cost = sum(int(r.get("total_cost") or 0) for r in dept_rows)

    # write totals to the summary block
    ws[f'{total_col}{first_label_row + 0}'].value = num_lecturers
    ws[f'{total_col}{first_label_row + 1}'].value = num_subjects
    ws[f'{total_col}{first_label_row + 2}'].value = lec_hours
    ws[f'{total_col}{first_label_row + 3}'].value = tut_hours
    ws[f'{total_col}{first_label_row + 4}'].value = prac_hours
    ws[f'{total_col}{first_label_row + 5}'].value = blend_hours
    ws[f'{total_col}{first_label_row + 6}'].value = total_cost

    # ---------- Charts ----------
    try:
        # Build mapping: {lecturer_name: total_cost}
        cost_by_lecturer = {}
        for r in dept_rows:
            name = r.get("lecturer_name", "Unknown")
            cost_by_lecturer[name] = cost_by_lecturer.get(name, 0) + int(r.get("total_cost") or 0)

        if cost_by_lecturer:
            # Insert data into a hidden block of cells (e.g., col M:N)
            start_row = 50  # safe row far below summary
            ws["M49"].value = "Lecturer"
            ws["N49"].value = "Cost"

            for i, (name, cost) in enumerate(cost_by_lecturer.items(), start=0):
                ws[f"M{start_row+i}"].value = name
                ws[f"N{start_row+i}"].value = cost

            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.title = "Cost by Lecturer"
            chart.y_axis.title = "RM"
            chart.x_axis.title = "Lecturer"

            data = Reference(ws, min_col=14, min_row=start_row, max_row=start_row+len(cost_by_lecturer)-1)  # col N=14
            cats = Reference(ws, min_col=13, min_row=start_row, max_row=start_row+len(cost_by_lecturer)-1)  # col M=13
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)
            chart.legend = None
            chart.width = 20
            chart.height = 10

            # Apply colors
            if chart.series:
                apply_colors(chart.series[0], len(cost_by_lecturer))

            # Add chart to sheet
            ws.add_chart(chart, "J7")
        
    except Exception as e:
        # If chart creation fails for any reason, don't break report generation
        logging.warning(f"Chart creation skipped: {e}")

def fill_department_sheet(ws, dept_code, dept_rows, start_date, end_date):
    """
    For a given dept sheet:
      - set Faculty/Centre (B3) and Date (B4)
      - write details (row 8..)
      - write the per-department summary (rows 12..18)
    """
    # Add image to worksheet
    ws.merge_cells('H2:H4')
    logo_path = os.path.join(current_app.root_path, 'static', 'img', 'Form INTI Logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 110
        img.height = 60
        img.anchor = 'H2'
        ws.add_image(img)

    # Header fields
    ws['B3'].value = dept_code
    ws['B4'].value = f"{format_date(start_date)} - {format_date(end_date)}"

    write_department_details(ws, dept_rows)
    write_department_summary(ws, dept_rows)  # auto-detects Total column

# Main generator
def generate_report_excel(report_type, start_date, end_date, report_details):
    try:
        # Load template
        base_dir = os.path.abspath(os.path.dirname(__file__))
        output_folder = os.path.join(base_dir, "temp")

        if report_type == "claim":
            template_name = "Claim Report - template.xlsx"
            output_filename = f"Claim Report_{format_file_date(start_date)} - {format_file_date(end_date)}.xlsx"
        else:  # default to requisition
            template_name = "Requisition Report - template.xlsx"
            output_filename = f"Requisition Report_{format_file_date(start_date)} - {format_file_date(end_date)}.xlsx"

        # Ensure output directory exists
        template_path = os.path.join(base_dir, "files", template_name)
        output_path = os.path.join(output_folder, output_filename)

        # Ensure output directory exists
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Workbook
        wb = load_workbook(template_path)
        ws = wb["Overall"]

        # Add image to worksheet
        ws.merge_cells('I2:I4')
        logo_path = os.path.join(current_app.root_path, 'static', 'img', 'Form INTI Logo.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 110
            img.height = 60
            img.anchor = 'I2'
            ws.add_image(img)
        
        # Date
        ws['B3'].value = f"{format_date(start_date)} - {format_date(end_date)}"

        # ===== Overall sheet =====
        summary_data_start = OVERALL_SUMMARY_DATA_START # Default summary start row

        # Insert Details
        n = len(report_details)
        if n > 1:
            # insert (n-1) rows after the first detail row to push summary down
            ws.insert_rows(OVERALL_DETAILS_START_ROW + 1, amount=n - 1)

            # Compute local anchors
            shift = max(0, n - 1)
            summary_data_start = OVERALL_SUMMARY_DATA_START + shift

        # style-copy + write each row
        for i, rec in enumerate(report_details):
            row_idx = OVERALL_DETAILS_START_ROW + i
            if i > 0:
                copy_row_style(ws, OVERALL_DETAILS_START_ROW, row_idx)
            insert_overall_details(ws, rec, row_idx)

        # merge equal departments in column A
        merge_same_department(ws, OVERALL_DETAILS_START_ROW, len(report_details), col="A")

        # Insert Summary
        write_overall_summary(ws, report_details, data_start_row=summary_data_start, put_chart=True)

        # ===== Department sheets =====
        bucket = group_details_by_department(report_details)

        for dept_code in DEPARTMENT_SHEETS:
            if dept_code not in wb.sheetnames:
                continue  # skip if template sheet missing
            ws = wb[dept_code]
            rows = bucket.get(dept_code, [])
            fill_department_sheet(ws, dept_code, rows, start_date, end_date)

        # Protect sheets
        for name in ["Overall"] + [s for s in DEPARTMENT_SHEETS if s in wb.sheetnames]:
            wb[name].protection.sheet = True

        wb.save(output_path)
        return output_path

    except Exception as e:
        logging.error(f"Error generating Excel file: {e}")
        raise
