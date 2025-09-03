import base64, io, logging, mimetypes, os, pytz, re, requests, tempfile
from app import app, db, mail
from app.database import handle_db_connection
from app.excel_generator import generate_claim_excel
from app.models import Admin, ClaimApproval, Department, Head, Lecturer, LecturerAttachment, LecturerClaim, LecturerSubject, Other, ProgramOfficer, Rate, RequisitionApproval, Subject 
from datetime import datetime
from flask import abort, current_app, jsonify, redirect, render_template, render_template_string, request, session, url_for
from flask_bcrypt import Bcrypt
from flask_mail import Message
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
from sqlalchemy import desc, func
bcrypt = Bcrypt()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.route('/lecturerHomepage', methods=['GET', 'POST'])
@handle_db_connection
def lecturerHomepage():
    if 'lecturer_id' not in session:
        return redirect(url_for('loginPage'))
    
    # Get distinct subject levels for this lecturer where requisition status is 'Completed'
    levels = (
        db.session.query(Subject.subject_level)
        .join(LecturerSubject, Subject.subject_id == LecturerSubject.subject_id)
        .join(RequisitionApproval, LecturerSubject.requisition_id == RequisitionApproval.approval_id)
        .filter(LecturerSubject.lecturer_id == session.get('lecturer_id'))
        .filter(RequisitionApproval.status == 'Completed')  # Only completed requisitions
        .distinct()
        .all()
    )
    levels = [level[0] for level in levels]  # Flatten the result from [(level1,), (level2,)] to [level1, level2]

    return render_template('lecturerHomepage.html', levels=levels)

@app.route('/get_subjects/<level>')
@handle_db_connection
def get_subjects(level):
    try:
        lecturer_id = session.get('lecturer_id')

        rows = (
            db.session.query(LecturerSubject, Subject, RequisitionApproval)
            .join(Subject, LecturerSubject.subject_id == Subject.subject_id)
            .join(RequisitionApproval, LecturerSubject.requisition_id == RequisitionApproval.approval_id)
            .filter(LecturerSubject.lecturer_id == lecturer_id)
            .filter(Subject.subject_level == level)
            .filter(RequisitionApproval.status == 'Completed')
            .all()
        )

        subject_list = []
        for ls, s, ra in rows:
            subject_list.append({
                'value': f"{s.subject_id}:{ls.requisition_id}",
                'label': f"{s.subject_code} - {s.subject_title} ({ls.start_date} → {ls.end_date})",
                'subject_code': s.subject_code,
                'subject_title': s.subject_title,
                'subject_id': s.subject_id,
                'requisition_id': ls.requisition_id,
                'start_date': ls.start_date.isoformat() if ls.start_date else '',
                'end_date': ls.end_date.isoformat() if ls.end_date else '',
            })

        return jsonify(success=True, subjects=subject_list)
    except Exception as e:
        current_app.logger.error(f"Error getting subjects by level: {e}")
        return jsonify(success=False, message=str(e), subjects=[])

@app.route('/get_subject_info')
@handle_db_connection
def get_subject_info():
    try:
        lecturer_id = session.get('lecturer_id')
        subject_id = request.args.get('subject_id', type=int)
        requisition_id = request.args.get('requisition_id', type=int)

        if not subject_id or not requisition_id:
            return jsonify(success=False, message="subject_id and requisition_id are required")

        ls = (
            db.session.query(LecturerSubject)
            .filter_by(lecturer_id=lecturer_id, subject_id=subject_id, requisition_id=requisition_id)
            .first()
        )
        if not ls:
            return jsonify(success=False, message="LecturerSubject not found")

        # Rate
        rate = Rate.query.get(ls.rate_id)
        hourly_rate = rate.amount if rate else 0

        # Claimed hours — IMPORTANT: filter by BOTH subject_id and requisition_id
        claimed = (
            db.session.query(
                func.sum(LecturerClaim.lecture_hours).label('claimed_lecture'),
                func.sum(LecturerClaim.tutorial_hours).label('claimed_tutorial'),
                func.sum(LecturerClaim.practical_hours).label('claimed_practical'),
                func.sum(LecturerClaim.blended_hours).label('claimed_blended')
            )
            .filter(
                LecturerClaim.lecturer_id == lecturer_id,
                LecturerClaim.subject_id == subject_id,
                LecturerClaim.requisition_id == requisition_id
            )
            .first()
        )

        claimed_lecture = claimed.claimed_lecture or 0
        claimed_tutorial = claimed.claimed_tutorial or 0
        claimed_practical = claimed.claimed_practical or 0
        claimed_blended = claimed.claimed_blended or 0

        return jsonify({
            'success': True,
            'start_date': ls.start_date.isoformat() if ls.start_date else '',
            'end_date': ls.end_date.isoformat() if ls.end_date else '',
            'unclaimed_lecture': max(0, (ls.total_lecture_hours or 0) - claimed_lecture),
            'unclaimed_tutorial': max(0, (ls.total_tutorial_hours or 0) - claimed_tutorial),
            'unclaimed_practical': max(0, (ls.total_practical_hours or 0) - claimed_practical),
            'unclaimed_blended': max(0, (ls.total_blended_hours or 0) - claimed_blended),
            'hourly_rate': hourly_rate,
            'rate_id': ls.rate_id or None
        })
    except Exception as e:
        current_app.logger.error(f"Error get_subject_info: {e}")
        return jsonify(success=False, message=str(e))

    
@app.route('/lecturerConversionResult', methods=['POST'])
@handle_db_connection
def lecturerConversionResult():
    if 'lecturer_id' not in session:
        return jsonify(success=False, error="Session expired. Please log in again."), 401

    try:
        # Debug: Print all form data
        print("Form Data:", request.form)
        
        # Get form data  
        lecturer = Lecturer.query.get(session.get('lecturer_id'))
        name = lecturer.name if lecturer else None
        department_code = lecturer.department.department_code if lecturer else None

        # Helper function to safely convert to int
        def safe_int(value, default=0):
            try:
                if isinstance(value, str):
                    return int(value.strip()) if value.strip() else default
                return int(value)
            except (ValueError, TypeError):
                return default

        subject_level = request.form.get('subject_level') 

        # Extract claim details from form
        claim_details = []
        i = 1
        while f'date{i}' in request.form:
            claim_details.append({
                'subject_id': int(request.form.get(f'subjectIdHidden{i}') or 0),
                'requisition_id': int(request.form.get(f'requisitionIdHidden{i}') or 0),
                'rate_id': int(request.form.get(f'rateIdHidden{i}') or 0),
                'subject_code': request.form.get(f'subjectCodeText{i}', ''),  # <— NEW
                'date': request.form.get(f'date{i}'),
                'lecture_hours': safe_int(request.form.get(f'lectureHours{i}'), 0),
                'tutorial_hours': safe_int(request.form.get(f'tutorialHours{i}'), 0),
                'practical_hours': safe_int(request.form.get(f'practicalHours{i}'), 0),
                'blended_hours': safe_int(request.form.get(f'blendedHours{i}'), 0),
                'remarks': request.form.get(f'remarks{i}')
            })
            i += 1

        if not claim_details:
            return jsonify(success=False, error="No claim details provided"), 400
               
        # Get department info
        department = Department.query.filter_by(department_code=department_code).first()
        department_id = department.department_id if department else None
        dean_name = department.dean_name if department else 'N/A'

        # Derive PO/Head from requisition_ids in the rows
        requisition_ids = {row['requisition_id'] for row in claim_details if row['requisition_id']}
        if not requisition_ids:
            return jsonify(success=False, error="No requisition linkage found in claim rows."), 400

        requisitions = (db.session.query(RequisitionApproval)
            .filter(RequisitionApproval.approval_id.in_(requisition_ids))
            .all())

        if not requisitions or len(requisitions) != len(requisition_ids):
            return jsonify(success=False, error="One or more requisitions not found."), 400

        # Enforce that all rows belong to the same PO/Head (simplest rule for a single approval)
        unique_po_ids = {r.po_id for r in requisitions if r.po_id is not None}
        unique_head_ids = {r.head_id for r in requisitions if r.head_id is not None}

        if len(unique_po_ids) != 1 or len(unique_head_ids) != 1:
            # You can choose to support mixed POs/Heads in one submission, but usually this is blocked.
            return jsonify(
                success=False,
                error="Selected rows span multiple Program Officers / Heads. "
                    "Please split into separate submissions by requisition."
            ), 400

        po_id = unique_po_ids.pop()
        head_id = unique_head_ids.pop()

        po = ProgramOfficer.query.get(po_id) if po_id else None
        head = Head.query.get(head_id) if head_id else None
        po_name = po.name if po else 'N/A'
        head_name = head.name if head else 'N/A'

        # Human Resources (excluding Ting Ting)
        hr = (Other.query
            .filter_by(role="Human Resources")
            .filter(Other.email != "tingting.eng@newinti.edu.my")
            .first())
        hr_name = hr.name if hr else 'N/A'

        # Validate required roles exist
        missing_roles = []
        if not po:   missing_roles.append("Program Officer")
        if not head: missing_roles.append("Head of Programme")
        if not department or not department.dean_name: missing_roles.append("Dean")
        if not hr:   missing_roles.append("Human Resources")
        if missing_roles:
            return jsonify(success=False, error=f"Missing required role(s): {', '.join(missing_roles)}"), 400

        # Generate Excel file
        output_path, sign_col = generate_claim_excel(
            name=name,
            department_code=department_code,
            subject_level=subject_level,
            claim_details=claim_details,
            po_name=po_name,
            head_name=head_name,
            dean_name=dean_name,
            hr_name=hr_name
        )

        file_name = os.path.basename(output_path)
        file_url, file_id = upload_to_drive(output_path, file_name)

        # Save to database
        approval = ClaimApproval(
            department_id=department_id,
            lecturer_id=session.get('lecturer_id'),
            po_id=po_id,
            head_id=head.head_id if head else None,
            subject_level=subject_level,
            sign_col=sign_col,
            file_id=file_id,
            file_name=file_name,
            file_url=file_url,
            status="Pending Acknowledgement by Lecturer",
            last_updated = get_current_datetime()
        )

        db.session.add(approval)
        db.session.flush()  # Get approval_id before committing
        
        approval_id = approval.approval_id

        # Add lecturer_claim entries with claim_id
        for claim_data in claim_details:
            # rate_id from row
            rate_id = claim_data['rate_id'] or None
            rate_amount = 0
            if rate_id:
                r = Rate.query.get(rate_id)
                rate_amount = r.amount if r else 0

            total_cost = rate_amount * (
                claim_data['lecture_hours'] + claim_data['tutorial_hours'] + claim_data['practical_hours'] + claim_data['blended_hours']
            )

            lc = LecturerClaim(
                lecturer_id=session.get('lecturer_id'),
                requisition_id=claim_data['requisition_id'],
                claim_id=approval_id,
                subject_id=claim_data['subject_id'],
                date=datetime.strptime(claim_data['date'], "%Y-%m-%d").date(),
                lecture_hours=claim_data['lecture_hours'],
                tutorial_hours=claim_data['tutorial_hours'],
                practical_hours=claim_data['practical_hours'],
                blended_hours=claim_data['blended_hours'],
                rate_id=rate_id,
                total_cost=total_cost
            )
            db.session.add(lc)

        db.session.commit()

        # ======= Handle Attachments for Lecturer ========
        attachments = request.files.getlist('upload_attachment')
        attachment_urls = []

        if attachments:
            drive_service = get_drive_service()
            lecturer_id = session.get('lecturer_id')
            if not lecturer_id:
                return jsonify({
                    'success': False,
                    'error': 'Lecturer session not found. Please login again.'
                }), 400

            for attachment in attachments:
                with tempfile.NamedTemporaryFile(delete=False) as tmp:
                    attachment.save(tmp.name)

                    file_metadata = {'name': attachment.filename}
                    media = MediaFileUpload(tmp.name, mimetype=attachment.mimetype, resumable=True)
                    uploaded = drive_service.files().create(
                        body=file_metadata,
                        media_body=media,
                        fields='id'
                    ).execute()

                    # Set public view permission
                    drive_service.permissions().create(
                        fileId=uploaded['id'],
                        body={'type': 'anyone', 'role': 'reader'},
                    ).execute()

                    file_url = f"https://drive.google.com/file/d/{uploaded['id']}/view"
                    attachment_urls.append((attachment.filename, file_url))
                    os.unlink(tmp.name)

            # Save to LecturerAttachment table
            for filename, url in attachment_urls:
                lecturer_attachment = LecturerAttachment(
                    attachment_name=filename,
                    attachment_url=url,
                    lecturer_id=lecturer_id,
                    claim_id=approval_id
                )
                db.session.add(lecturer_attachment)

            db.session.commit()

        return jsonify({
            'success': True,
            'file_url': file_url,
            'attachments': [{'name': fn, 'url': url} for fn, url in attachment_urls]
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in result route: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route('/lecturerConversionResultPage')
def lecturerConversionResultPage():
    if 'lecturer_id' not in session:
        return redirect(url_for('loginPage'))
    
    approval = ClaimApproval.query.filter_by(lecturer_id=session.get('lecturer_id')).order_by(ClaimApproval.approval_id.desc()).first()
    return render_template('lecturerConversionResultPage.html', file_url=approval.file_url)

@app.route('/lecturerRecordsPage')
@handle_db_connection
def lecturerRecordsPage():
    if 'lecturer_id' not in session:
        return redirect(url_for('loginPage'))
    
    # Set default tab if none exists
    if 'lecturerRecordsPage_currentTab' not in session:
        session['lecturerRecordsPage_currentTab'] = 'claimDetails'

    lecturer_id = session['lecturer_id']

    # Get all lecturer_subject records for the lecturer
    subjects = (
        db.session.query(
            LecturerSubject,
            Subject.subject_code,
            Subject.subject_title,
            Subject.subject_level
        )
        .join(Subject, LecturerSubject.subject_id == Subject.subject_id)
        .join(RequisitionApproval, LecturerSubject.requisition_id == RequisitionApproval.approval_id)
        .filter(LecturerSubject.lecturer_id == lecturer_id)
        .filter(RequisitionApproval.status == 'Completed')  # Only completed requisitions
        .order_by(desc(RequisitionApproval.approval_id))   # Descending order
        .all()
    )

    claimDetails = []

    for ls, code, title, level in subjects:
        # Sum all claimed hours for this subject by this lecturer
        claimed = (
            db.session.query(
                func.coalesce(func.sum(LecturerClaim.lecture_hours), 0),
                func.coalesce(func.sum(LecturerClaim.tutorial_hours), 0),
                func.coalesce(func.sum(LecturerClaim.practical_hours), 0),
                func.coalesce(func.sum(LecturerClaim.blended_hours), 0)
            )
            .filter_by(lecturer_id=lecturer_id, subject_id=ls.subject_id)
            .first()
        )

        remaining = {
            'subject_code': code,
            'subject_title': title,
            'subject_level': level,
            'start_date': ls.start_date,
            'end_date': ls.end_date,
            'hourly_rate': ls.rate.amount if ls.rate else 0,
            'lecture_hours': ls.total_lecture_hours - claimed[0],
            'tutorial_hours': ls.total_tutorial_hours - claimed[1],
            'practical_hours': ls.total_practical_hours - claimed[2],
            'blended_hours': ls.total_blended_hours - claimed[3],
        }

        claimDetails.append(remaining)
   
    lecturerAttachments = LecturerAttachment.query.all()

    return render_template('lecturerRecordsPage.html', 
                           claimDetails=claimDetails, 
                           lecturerAttachments=lecturerAttachments)

@app.route('/set_lecturerRecordsPage_tab', methods=['POST'])
def set_lecturerRecordsPage_tab():
    if 'lecturer_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    session['lecturerRecordsPage_currentTab'] = data.get('lecturerRecordsPage_currentTab')
    return jsonify({'success': True})

@app.route('/lecturerApprovalsPage')
@handle_db_connection
def lecturerApprovalsPage():
    if 'lecturer_id' not in session:
        return redirect(url_for('loginPage'))

    approvals = ClaimApproval.query.filter_by(lecturer_id=session.get('lecturer_id')).all()
    return render_template('lecturerApprovalsPage.html', approvals=approvals)

@app.route('/check_claim_status/<int:approval_id>')
def check_claim_status(approval_id):
    approval = ClaimApproval.query.get_or_404(approval_id)
    return jsonify({'status': approval.status})

@app.route('/api/lecturer_review_claim/<approval_id>', methods=['POST'])
@handle_db_connection
def lecturer_review_claim(approval_id):
    try:
        data = request.get_json()
        signature_data = data.get("image")

        if not signature_data or "," not in signature_data:
            return jsonify(success=False, error="Invalid image data format")

        # Fetch approval record
        approval = ClaimApproval.query.get(approval_id)
        if not approval:
            return jsonify(success=False, error="Approval record not found")

        # Process signature and upload updated file
        process_signature_and_upload(approval, signature_data, "A")

        # Update status after signature inserted
        approval.status = "Pending Acknowledgement by PO"
        approval.last_updated = get_current_datetime()
        db.session.commit()

        try:
            notify_approval(approval, approval.program_officer.email if approval.program_officer else None, "po_review_claim", "Program Officer")
        except Exception as e:
            logging.error(f"Failed to notify PO: {e}")    

        return jsonify(success=True)

    except Exception as e:
        logging.error(f"Error uploading signature: {e}")
        return jsonify(success=False, error=str(e)), 500
    
@app.route('/api/po_review_claim/<approval_id>', methods=['GET', 'POST'])
def po_review_claim(approval_id):
    approval = ClaimApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    if request.method == 'GET':
        if is_already_reviewed(approval, ["Rejected by PO", "Pending Acknowledgement by HOP"]):
            return render_template_string(f"""
                <h2 style="text-align: center; color: red;">This request has already been reviewed.</h2>
                <p style="text-align: center;">Status: {approval.status}</p>
            """)
        return render_template("reviewClaimApprovalRequest.html", approval=approval)

    # POST logic
    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "B")
            approval.status = "Pending Acknowledgement by HOP"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            try:
                notify_approval(approval, approval.head.email if approval.head else None, "head_review_claim", "Head of Programme")
            except Exception as e:
                logging.error(f"Failed to notify HOP: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by PO - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_claim_and_attachments(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("PO", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")

        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400

@app.route('/api/head_review_claim/<approval_id>', methods=['GET', 'POST'])
def head_review_claim(approval_id):
    approval = ClaimApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    if request.method == 'GET':
        if is_already_reviewed(approval, ["Rejected by HOP", "Pending Acknowledgement by Dean / HOS"]):
            return render_template_string(f"""
                <h2 style="text-align: center; color: red;">This request has already been reviewed.</h2>
                <p style="text-align: center;">Status: {approval.status}</p>
            """)
        return render_template("reviewClaimApprovalRequest.html", approval=approval)

    # POST logic
    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "C")
            approval.status = "Pending Acknowledgement by Dean / HOS"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            try:
                notify_approval(approval, approval.department.dean_email if approval.department else None, "dean_review_claim", "Dean / HOS")
            except Exception as e:
                logging.error(f"Failed to notify Dean: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by HOP - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_claim_and_attachments(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("HOP", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")

        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400

@app.route('/api/dean_review_claim/<approval_id>', methods=['GET', 'POST'])
def dean_review_claim(approval_id):
    approval = ClaimApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    if request.method == 'GET':
        if is_already_reviewed(approval, ["Rejected by Dean / HOS", "Pending Acknowledgement by HR"]):
            return render_template_string(f"""
                <h2 style="text-align: center; color: red;">This request has already been reviewed.</h2>
                <p style="text-align: center;">Status: {approval.status}</p>
            """)
        return render_template("reviewClaimApprovalRequest.html", approval=approval)
    
    # POST logic
    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "E")
            approval.status = "Pending Acknowledgement by HR"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            hr = Other.query.filter(Other.role == "Human Resources", Other.email != "tingting.eng@newinti.edu.my").first()

            try:
                notify_approval(approval, hr.email if hr else None, "hr_review_claim", "HR")
            except Exception as e:
                logging.error(f"Failed to notify HR: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by Dean / HOS - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_claim_and_attachments(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("Dean", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")

        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400

@app.route('/api/hr_review_claim/<approval_id>', methods=['GET', 'POST'])
def hr_review_claim(approval_id):
    approval = ClaimApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    if request.method == 'GET':
        if is_already_reviewed(approval, ["Rejected by HR", "Completed"]):
            return render_template_string(f"""
                <h2 style="text-align: center; color: red;">This request has already been reviewed.</h2>
                <p style="text-align: center;">Status: {approval.status}</p>
            """)
        return render_template("reviewClaimApprovalRequest.html", approval=approval)

    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "G")
            approval.status = "Completed"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            try:
                subject = f"Part-time Lecturer Claim Approval Request Completed  - {approval.lecturer.name} ({approval.subject_level})"
                body = (
                    f"Dear All,\n\n"
                    f"The part-time lecturer claim request has been fully approved by all parties.\n\n"
                    f"Please click the link below to access the final approved file:\n"
                    f"{approval.file_url}\n\n"
                    "Thank you for your cooperation.\n\n"
                    "Best regards,\n"
                    "The CourseXcel Team"
                )
                
                final_hr_email = "tingting.eng@newinti.edu.my"
                admin = Admin.query.filter_by(admin_id=1).first()

                # Base recipients from related models
                recipients = [
                    approval.lecturer.email if approval.lecturer else None,
                    approval.program_officer.email if approval.program_officer else None,
                    approval.head.email if approval.head else None,
                    approval.department.dean_email if approval.department else None,
                ]

                # Get "Other" roles
                hr = Other.query.filter_by(role="Human Resources").filter(Other.email != final_hr_email).first()

                # Append first HR if exists
                if hr and hr.email:
                    recipients.append(hr.email)

                # Append final HR and admin
                recipients.append(final_hr_email)
                if admin and admin.email:
                    recipients.append(admin.email)

                # Filter out any Nones or duplicates
                recipients = list(filter(None, set(recipients)))

                send_email(recipients, subject, body)
            except Exception as e:
                logging.error(f"Failed to notify All: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by HR - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_claim_and_attachments(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("HR", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")
        
        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400

@app.route('/api/void_claim/<approval_id>', methods=['POST'])
@handle_db_connection
def void_claim(approval_id):
    try:
        data = request.get_json()
        reason = data.get("reason", "").strip()

        if not reason:
            return jsonify(success=False, error="Reason for voiding is required."), 400

        approval = ClaimApproval.query.get(approval_id)
        if not approval:
            return jsonify(success=False, error="Approval record not found."), 404

        current_status = approval.status

        # Allow voiding only at specific stages
        if current_status in [
            "Pending Acknowledgement by Lecturer",
            "Pending Acknowledgement by PO",
            "Pending Acknowledgement by HOP",
            "Pending Acknowledgement by Dean / HOS",
            "Pending Acknowledgement by HR"
        ]:
            approval.status = f"Voided - {reason}"
            approval.last_updated = get_current_datetime()

            # Delete related records and rename approval file
            delete_claim_and_attachments(approval.approval_id, "VOIDED")
     
        else:
            return jsonify(success=False, error="Request cannot be voided at this stage."), 400

        # Determine recipients based on current stage
        recipients = []
        if current_status == "Pending Acknowledgement by HOP":
            recipients = [approval.program_officer.email]
        elif current_status == "Pending Acknowledgement by Dean / HOS":
            recipients = [approval.program_officer.email, approval.head.email]
        elif current_status == "Pending Acknowledgement by HR":
            recipients = [approval.program_officer.email, approval.head.email, approval.department.dean_email]

        recipients = list(set(filter(None, recipients)))  # Remove duplicates and None

        # Send notification emails
        subject = f"Part-time Lecturer Claim Request Voided - {approval.lecturer.name} ({approval.subject_level})"
        body = (
            f"Dear All,\n\n"
            f"The part-time lecturer claim request has been voided by the Requester.\n"
            f"Reason: {reason}\n\n"
            f"Please review the file here:\n{approval.file_url}\n"
            "Please do not take any further action on this request.\n\n"
            "Thank you,\n"
            "The CourseXcel Team"
        )

        if recipients:
            success = send_email(recipients, subject, body)
            if not success:
                logging.error(f"Failed to send void notification email to: {recipients}")
        
        return jsonify(success=True)

    except Exception as e:
        logging.error(f"Error voiding claim: {e}")
        return jsonify(success=False, error="Internal server error."), 500

@app.route('/lecturerProfilePage')
def lecturerProfilePage():
    lecturer_email = session.get('lecturer_email')
    if not lecturer_email:
        return redirect(url_for('loginPage'))

    lecturer = Lecturer.query.filter_by(email=lecturer_email).first()

    return render_template('lecturerProfilePage.html', lecturer=lecturer)

def get_drive_service():
    SERVICE_ACCOUNT_FILE = '/home/TomazHayden/coursexcel-459515-3d151d92b61f.json'
    SCOPES = ['https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

def upload_to_drive(file_path, file_name):
    try:
        drive_service = get_drive_service()

        file_metadata = {
            'name': file_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet'  # Convert to Google Sheets
        }

        media = MediaFileUpload(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )

        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        # Make file publicly accessible
        drive_service.permissions().create(
            fileId=file.get('id'),
            body={'type': 'anyone', 'role': 'reader'},
        ).execute()

        file_id = file.get('id')
        file_url = f"https://docs.google.com/spreadsheets/d/{file_id}/edit"

        return file_url, file_id

    except Exception as e:
        logging.error(f"Failed to upload to Google Drive: {e}")
        raise

def download_from_drive(file_id):
    drive_service = get_drive_service()

    request = drive_service.files().export_media(fileId=file_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    output_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), "temp")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    local_path = os.path.join(output_folder, f"{file_id}.xlsx")

    fh = io.FileIO(local_path, 'wb')
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    fh.close()
    return local_path

def get_current_datetime():
    malaysia_tz = pytz.timezone('Asia/Kuala_Lumpur')
    now_myt = datetime.now(malaysia_tz)
    return now_myt.strftime('%a, %d %b %y, %I:%M:%S %p')
          
def save_signature_image(signature_data, approval_id, temp_folder):
    try:
        header, encoded = signature_data.split(",", 1)
        binary_data = base64.b64decode(encoded)
        image = Image.open(BytesIO(binary_data))
        temp_image_path = os.path.join(temp_folder, f"signature_{approval_id}.png")
        image.save(temp_image_path)
        return temp_image_path
    except Exception as e:
        logging.error(f"Signature decoding error: {e}")
        return None

def insert_signature_and_date(local_excel_path, signature_path, cell_prefix, row, updated_path):
    wb = load_workbook(local_excel_path)
    ws = wb['Sheet1']

    # Insert signature
    sign_cell = f"{cell_prefix}{row}"
    signature_img = ExcelImage(signature_path)
    signature_img.width = 100
    signature_img.height = 30
    ws.add_image(signature_img, sign_cell)

    # Insert date
    date_cell = f"{cell_prefix}{row + 3}"
    malaysia_time = datetime.now(pytz.timezone('Asia/Kuala_Lumpur'))
    ws[date_cell] = f"Date: {malaysia_time.strftime('%d/%m/%Y')}"

    wb.save(updated_path)

def process_signature_and_upload(approval, signature_data, col_letter):
    temp_folder = os.path.join("temp")
    os.makedirs(temp_folder, exist_ok=True)

    temp_image_path = save_signature_image(signature_data, approval.approval_id, temp_folder)
    if not temp_image_path:
        raise ValueError("Invalid signature image data")

    local_excel_path = download_from_drive(approval.file_id)
    updated_excel_path = os.path.join(temp_folder, approval.file_name)

    try:
        insert_signature_and_date(local_excel_path, temp_image_path, col_letter, approval.sign_col, updated_excel_path)
        new_file_url, new_file_id = upload_to_drive(updated_excel_path, approval.file_name)

        # Delete old versions from Drive except the new one
        drive_service = get_drive_service()

        # Delete old file by stored file ID (approval.file_id) if different from new_file_id
        if approval.file_id and approval.file_id != new_file_id:
            try:
                drive_service.files().delete(fileId=approval.file_id).execute()
                logging.info(f"Deleted old file with ID {approval.file_id}")
            except Exception as e:
                logging.warning(f"Failed to delete old file {approval.file_id}: {e}")
                
        # Update DB record
        approval.file_url = new_file_url
        approval.file_id = new_file_id
        approval.last_updated = get_current_datetime()
        db.session.commit()

    finally:
        # Clean up temp files safely even if an exception occurs
        for path in [temp_image_path, local_excel_path, updated_excel_path]:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception as e:
                logging.warning(f"Failed to remove temp file {path}: {e}")

def delete_claim_and_attachments(approval_id, suffix):
    # Fetch the approval record first
    approval = ClaimApproval.query.get(approval_id)
    if not approval:
        logging.warning(f"No approval record found for ID {approval_id}")
        return
    
    # Rename file
    if approval.file_name:
        name, ext = os.path.splitext(approval.file_name)
        new_file_name = f"{name}_{suffix}{ext}"

        # Update Google Drive file name
        if approval.file_id:
            try:
                drive_service = get_drive_service()
                file_metadata = {"name": new_file_name}
                drive_service.files().update(fileId=approval.file_id, body=file_metadata).execute()
                logging.info(f"Renamed Google Drive file {approval.file_name} -> {new_file_name}")
            except Exception as e:
                logging.error(f"Failed to rename Google Drive file '{approval.file_name}': {e}")

        # Update DB field
        approval.file_name = new_file_name
    
    # Delete linked LecturerClaim entries
    LecturerClaim.query.filter_by(claim_id=approval_id).delete(synchronize_session=False)

    # Delete related attachments
    try:
        drive_service = get_drive_service()
        attachments_to_delete = LecturerAttachment.query.filter_by(claim_id=approval_id).all()
        for attachment_record in attachments_to_delete:
            try:
                # Extract file ID from Google Drive URL
                match = re.search(r'/d/([a-zA-Z0-9_-]+)', attachment_record.attachment_url)
                if not match:
                    logging.warning(f"Invalid Google Drive URL format for attachment {attachment_record.attachment_name}")
                    continue
                drive_attachment_id = match.group(1)

                # Delete file from Google Drive
                drive_service.files().delete(fileId=drive_attachment_id).execute()

                # Delete attachment record from database
                db.session.delete(attachment_record)

            except Exception as e:
                logging.error(f"Failed to delete Drive attachment '{attachment_record.attachment_name}': {e}")
    except Exception as e:
        logging.error(f"Failed to initialize Drive service or delete attachments: {e}")

    # Commit DB changes
    db.session.commit()

def is_already_voided(approval):
    if "Voided" in approval.status:
        return render_template_string(f"""
            <h2 style="text-align: center; color: red;">This request has already been voided.</h2>
            <p style="text-align: center;">Status: {approval.status}</p>
        """)
    return None

def is_already_reviewed(approval, expected_statuses):
    return any(status in approval.status for status in expected_statuses)

def get_attachments_for_approval(approval_id):
    # Returns a list of LecturerAttachment objects
    return LecturerAttachment.query.filter_by(claim_id=approval_id).all()

def send_email(recipients, subject, body, attachments=None):
    # attachments: list of dicts, each dict has keys: 'filename' and 'url'
    
    try:
        # Ensure recipients is always a list
        if isinstance(recipients, str):
            recipients = [recipients]

        msg = Message(subject, recipients=recipients, body=body)

        # Attach files
        if attachments:
            for att in attachments:
                try:
                    url = att['url']
                    filename = att['filename']

                    # If Google Drive link, convert to direct-download
                    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
                    if m:
                        file_id = m.group(1)
                        url = f"https://drive.google.com/uc?export=download&id={file_id}"

                    resp = requests.get(url, allow_redirects=True, timeout=30)
                    resp.raise_for_status()

                    msg.attach(filename, "application/pdf", resp.content)
                except Exception as e:
                    app.logger.error(f"Failed to attach {att.get('filename')}: {e}")

        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Failed to send email: {e}")
        return False
    
def notify_approval(approval, recipient_email, next_review_route, greeting):
    if not recipient_email:
        logging.error("No recipient email provided for approval notification.")
        return

    review_url = url_for(next_review_route, approval_id=approval.approval_id, _external=True)

    # Get all attachments for this approval
    attachments = get_attachments_for_approval(approval.approval_id)

    # Convert attachments to list of dicts with filename & URL
    attachment_list = [
        {'filename': att.attachment_name, 'url': att.attachment_url}
        for att in attachments
    ]

    subject = f"Part-time Lecturer Claim Approval Request - {approval.lecturer.name} ({approval.subject_level})"
    body = (
        f"Dear {greeting},\n\n"
        f"There is a part-time lecturer claim request pending your review and approval.\n\n"
        f"Please click the link below to review and approve or reject the request:\n"
        f"{review_url}\n\n"
        "Attachments are included for your reference.\n\n"
        "Thank you,\n"
        "The CourseXcel Team"
    )

    send_email(recipient_email, subject, body, attachments=attachment_list)

def send_rejection_email(role, approval, reason):
    subject = f"Part-time Lecturer Claim Request Rejected - {approval.lecturer.name} ({approval.subject_level})"

    role_names = {
        "PO": "Program Officer",
        "HOP": "Head of Programme",
        "Dean": "Dean / Head of School",
        "HR": "HR"
    }

    recipients_map = {
        "PO": [approval.lecturer.email] if approval.lecturer else [],
        "HOP": [
            approval.lecturer.email if approval.lecturer else None,
            approval.program_officer.email if approval.program_officer else None
        ],
        "Dean": [
            approval.lecturer.email if approval.lecturer else None,
            approval.program_officer.email if approval.program_officer else None,
            approval.head.email if approval.head else None
        ],
        "HR": [
            approval.lecturer.email if approval.lecturer else None,
            approval.program_officer.email if approval.program_officer else None,
            approval.head.email if approval.head else None,
            approval.department.dean_email if approval.department else None
        ]
    }

     # Clean up None values and deduplicate
    recipients = list(set(filter(None, recipients_map.get(role, []))))

    rejected_by = role_names.get(role, "Unknown Role")
    greeting = "Dear Requester" if role == "PO" else "Dear All"

    body = (
        f"{greeting},\n\n"
        f"The part-time lecturer claim approval request has been rejected by the {rejected_by}.\n\n"
        f"Reason for rejection: {reason}\n\n"
        f"You can review the file here:\n{approval.file_url}\n\n"
        "Please take necessary actions.\n\n"
        "Thank you,\n"
        "The CourseXcel Team"
    )

    send_email(recipients, subject, body)
