import base64, io, logging, os, pytz, re, requests
from app import app, db, mail
from app.auth import login_user
from app.database import handle_db_connection
from app.models import Admin, Department, Head, Lecturer, LecturerSubject, Other, ProgramOfficer, Rate, RequisitionApproval, RequisitionAttachment, Subject 
from datetime import datetime, timezone
from flask import flash, jsonify, redirect, render_template, render_template_string, request, session, url_for
from flask_bcrypt import Bcrypt
from flask_mail import Message
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from io import BytesIO
from itsdangerous import URLSafeTimedSerializer
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
from sqlalchemy.exc import IntegrityError

bcrypt = Bcrypt()
logger = logging.getLogger(__name__)

# ============================================================
#  Login & Logout Functions
# ============================================================
@app.route('/')
def index():
    return redirect(url_for('loginPage'))

@app.route('/loginPage', methods=['GET', 'POST'])
def loginPage():    
    locked = False

    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        # Track login attempts
        attempts = session.get('login_attempts', {})
        user_attempt = attempts.get(email, {"count": 0})

        # If already locked, force forgot password
        if user_attempt["count"] >= 3:
            locked = True
            return render_template(
                'loginPage.html',
                error_message="Account locked due to too many failed attempts.\nPlease use Forgot Password to reset.",
                locked=locked
            )

        # Normal login
        role = login_user(email, password)

        if role:  # login success
            # Reset attempts on success
            if email in attempts:
                del attempts[email]
            session['login_attempts'] = attempts

            if role == 'admin':
                return redirect(url_for('adminHomepage'))
            elif role == 'program_officer':
                return redirect(url_for('poHomepage'))
            elif role == 'lecturer':
                return redirect(url_for('lecturerHomepage'))
        
        else:  # login failed
            user_attempt["count"] += 1
            attempts[email] = user_attempt
            session['login_attempts'] = attempts

            if user_attempt["count"] >= 3:
                locked = True
                return render_template(
                    'loginPage.html',
                    error_message="Account locked after 3 failed attempts.\nPlease use Forgot Password to reset.",
                    locked=locked
                )

            return render_template(
                'loginPage.html',
                error_message=f"Invalid email or password. Attempt {user_attempt['count']} of 3.",
                locked=False
            )
        
    return render_template('loginPage.html', locked=False)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('loginPage'))

# ============================================================
#  Password-Related Functions
# ============================================================
@app.route('/api/forgot_password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    email = data.get('email')

    if not email:
        return jsonify({'success': False, 'message': 'Email is required.'})

    # Check if user exists in any model
    user = Admin.query.filter_by(email=email).first() \
        or ProgramOfficer.query.filter_by(email=email).first() \
        or Lecturer.query.filter_by(email=email).first()

    if not user:
        return jsonify({'success': False, 'message': 'User not found. Please ensure you insert the correct email.'}), 404

    s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    token = s.dumps(email, salt='reset-password')
    reset_url = url_for('reset_password', token=token, _external=True)

    msg = Message(f'CourseXcel - Password Reset Request', recipients=[email])
    msg.body = f'''Hi,

We received a request to reset your password for your CourseXcel account.

To reset your password, click the link below:
{reset_url}

If you did not request this, you can safely ignore this email.

Thank you,
The CourseXcel Team
'''
    mail.send(msg)

    return jsonify({'success': True, 'message': 'Reset link sent to your email.'})

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = s.loads(token, salt='reset-password', max_age=3600)
    except Exception:
        return 'The reset link is invalid or has expired.', 400

    # Role detection logic
    for model in [Admin, ProgramOfficer, Lecturer]:
        user = model.query.filter_by(email=email).first()
        if user:
            break
    else:
        return 'Role could not be identified.', 400

    if request.method == 'POST':
        new_password = request.form.get('new_password')
        hashed_password = bcrypt.generate_password_hash(new_password).decode('utf-8')
        user.password = hashed_password
        db.session.commit()

        # Unlock user by clearing login attempts
        attempts = session.get('login_attempts', {})
        if email in attempts:
            del attempts[email]
            session['login_attempts'] = attempts

        flash("Password has been reset successfully.\nPlease log in with your new password.", "success")
        return redirect(url_for('loginPage'))

    html_content = '''
        <link href="https://fonts.googleapis.com/css2?family=Roboto&display=swap" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.4/css/all.min.css">
        <style>
            html, body {
                height: 100%;
                overflow: hidden;  /* Prevent page scroll */
            }
            body {
                font-family: 'Roboto', sans-serif;
                padding: 2rem;
                display: flex;
                justify-content: center;
                align-items: center;
                min-height: 100vh;
                background-color: #f9f9f9;
            }
            form {
                width: 100%;
                max-width: 400px;
                background: white;
                padding: 2rem;
                border-radius: 8px;
                box-shadow: 0 0 10px rgba(0,0,0,0.1);
            }
            h2 {
                text-align: center;
                margin-bottom: 30px;
            }
            .input-group {
                position: relative;
                margin-bottom: 25px;
            }
            .input-group input {
                width: 100%;
                padding: 10px 40px 10px 0;
                font-size: 16px;
                border: none;
                border-bottom: 1px solid #ddd;
                background: transparent;
                outline: none;
            }
            .input-group label {
                position: absolute;
                top: -10px;
                left: 0;
                font-size: 12px;
                color: #777;
            }
            .input-group button {
                position: absolute;
                right: 0;
                top: 50%;
                transform: translateY(-50%);
                border: none;
                background: none;
                cursor: pointer;
                font-size: 16px;
            }
            .reset-btn {
                width: 100%;
                padding: 12px;
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                cursor: pointer;
                display: block;
                margin-top: 10px;
            }
        </style>

        <form method="post" onsubmit="return validatePasswords()">
            <h2>Reset Password</h2>

            <div class="input-group">
                <input type="password" name="new_password" id="new_password" required>
                <label for="new_password">New Password</label>
                <button type="button" onclick="togglePassword('new_password', this)">
                    <i class="fas fa-eye"></i>
                </button>
            </div>    

            <div class="input-group">
                <input type="password" name="confirm_password" id="confirm_password" required>
                <label for="confirm_password">Confirm Password</label>
                <button type="button" onclick="togglePassword('confirm_password', this)">
                    <i class="fas fa-eye"></i>
                </button>
            </div>    

            <button class="reset-btn" type="submit">Confirm</button>
        </form>

        <script>
        function togglePassword(inputId, button) {
            var input = document.getElementById(inputId);
            var icon = button.querySelector('i');

            if (input.type === 'password') {
                input.type = 'text';
                icon.classList.replace('fa-eye', 'fa-eye-slash');
            } else {
                input.type = 'password';
                icon.classList.replace('fa-eye-slash', 'fa-eye');
            }
        }

        function validatePasswords() {
            const newPassword = document.getElementById('new_password').value;
            const confirmPassword = document.getElementById('confirm_password').value;

            if (newPassword !== confirmPassword) {
                alert('Passwords do not match.');
                return false;
            }
            
            const minLength = 8;
            const hasLetter = /[A-Za-z]/.test(newPassword);
            const hasNumber = /[0-9]/.test(newPassword);
            const hasSpecial = /[!@#$%^&*(),.?":{}|<>_]/.test(newPassword);

            if (newPassword.length < minLength) {
                alert('Password must be at least 8 characters long.');
                return false;
            }
            if (!hasLetter || !hasNumber || !hasSpecial) {
                alert('Password must include letters, numbers, and special symbols.');
                return false;
            }

            return true;
        }
        </script>
        '''

    return render_template_string(html_content)
    
@app.route('/api/change_password', methods=['POST'])
@handle_db_connection
def change_password():
    try:
        data = request.get_json()
        new_password = data.get('new_password')
        role = data.get('role')

        if not new_password or not role:
            return jsonify({'success': False, 'message': 'New password and role are required.'})

        # Map role to model and session key
        role_config = {
            'admin': (Admin, 'admin_email'),
            'program_officer': (ProgramOfficer, 'po_email'),
            'lecturer': (Lecturer, 'lecturer_email')
        }

        if role not in role_config:
            return jsonify({'success': False, 'message': 'Invalid role.'})

        Model, session_key = role_config[role]
        email = session.get(session_key)

        if not email:
            return jsonify({'success': False, 'message': 'Not logged in.'})

        user = Model.query.filter_by(email=email).first()
        if not user:
            return jsonify({'success': False, 'message': f'{role.replace("_", " ").title()} not found.'})

        hashed_password = bcrypt.generate_password_hash(new_password).decode('utf-8')
        user.password = hashed_password
        db.session.commit()

        return jsonify({'success': True, 'message': 'Password changed successfully.'})

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error while changing password: {e}")
        return jsonify({'success': False, 'message': str(e)}) 

# ============================================================
#  Utility Functions
# ============================================================
def get_current_utc():
    """Return current UTC time as a datetime object (for DB)."""
    return datetime.now(timezone.utc)

def format_utc(dt):
    """Convert a datetime object to Malaysia timezone and format as string."""
    if not dt:
        return None
    malaysia_tz = pytz.timezone('Asia/Kuala_Lumpur')
    return dt.astimezone(malaysia_tz).strftime('%a, %d %b %y, %I:%M:%S %p')

def to_utc_aware(dt):
    """Ensure datetime is timezone-aware in UTC. If naive, assume UTC and set tzinfo."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

def get_drive_service():
    SERVICE_ACCOUNT_FILE = '/home/TomazHayden/coursexcel-459515-3d151d92b61f.json'
    SCOPES = ['https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

# ============================================================
#  Google Drive Operations
# ============================================================
def download_from_drive(file_id):
    logger.info(f"Downloading file from Drive: {file_id}")
    drive_service = get_drive_service()

    request = drive_service.files().export_media(
        fileId=file_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    output_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), "temp")
    os.makedirs(output_folder, exist_ok=True)
    local_path = os.path.join(output_folder, f"{file_id}.xlsx")

    fh = io.FileIO(local_path, 'wb')
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()
        logger.debug(f"Download progress: {int(status.progress() * 100)}%")

    fh.close()
    logger.info(f"File downloaded successfully: {local_path}")
    return local_path

def upload_to_drive(file_path, file_name):
    logger.info(f"Uploading file to Drive: {file_name}")
    try:
        drive_service = get_drive_service()

        file_metadata = {
            'name': file_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet'  # Convert to Google Sheets
        }

        media = MediaFileUpload(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )

        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        # Make file publicly accessible
        drive_service.permissions().create(
            fileId=file.get('id'),
            body={'type': 'anyone', 'role': 'reader'},
        ).execute()

        file_id = file.get('id')
        file_url = f"https://docs.google.com/spreadsheets/d/{file_id}/edit"
        logger.info(f"File uploaded successfully. ID: {file_id}")

        return file_url, file_id

    except Exception as e:
        logger.error(f"Failed to upload to Google Drive: {e}")
        raise

# ============================================================
#  Signature Processing
# ============================================================
def save_signature_image(signature_data, approval_id, temp_folder):
    try:
        header, encoded = signature_data.split(",", 1)
        binary_data = base64.b64decode(encoded)
        image = Image.open(BytesIO(binary_data))
        temp_image_path = os.path.join(temp_folder, f"signature_{approval_id}.png")
        image.save(temp_image_path)
        logger.debug(f"Signature image saved: {temp_image_path}")
        return temp_image_path
    except Exception as e:
        logger.error(f"Signature decoding error: {e}")
        return None

def insert_signature_and_date(local_excel_path, signature_path, cell_prefix, row, updated_path):
    logger.info(f"Inserting signature and date at row {row}")
    wb = load_workbook(local_excel_path)
    ws = wb.active

    # Insert signature
    sign_cell = f"{cell_prefix}{row}"
    signature_img = ExcelImage(signature_path)
    signature_img.width = 100
    signature_img.height = 30
    ws.add_image(signature_img, sign_cell)

    # Insert date
    date_cell = f"{cell_prefix}{row + 3}"
    malaysia_time = datetime.now(pytz.timezone('Asia/Kuala_Lumpur'))
    ws[date_cell] = f"Date: {malaysia_time.strftime('%d/%m/%Y')}"

    wb.save(updated_path)
    logger.info(f"Signature and date inserted, file saved: {updated_path}")

def process_signature_and_upload(approval, signature_data, col_letter):
    logger.info(f"Processing signature for approval ID: {approval.approval_id}")
    temp_folder = os.path.join("temp")
    os.makedirs(temp_folder, exist_ok=True)

    temp_image_path = save_signature_image(signature_data, approval.approval_id, temp_folder)
    if not temp_image_path:
        raise ValueError("Invalid signature image data")

    local_excel_path = download_from_drive(approval.file_id)
    updated_excel_path = os.path.join(temp_folder, approval.file_name)

    try:
        insert_signature_and_date(local_excel_path, temp_image_path, col_letter, approval.sign_col, updated_excel_path)
        new_file_url, new_file_id = upload_to_drive(updated_excel_path, approval.file_name)

        # Delete old file if needed
        drive_service = get_drive_service()
        if approval.file_id and approval.file_id != new_file_id:
            try:
                drive_service.files().delete(fileId=approval.file_id).execute()
                logger.info(f"Deleted old Drive file: {approval.file_id}")
            except Exception as e:
                logger.warning(f"Failed to delete old file {approval.file_id}: {e}")
                
        # Update DB record
        approval.file_url = new_file_url
        approval.file_id = new_file_id
        approval.last_updated = get_current_utc()
        db.session.commit()

    finally:
        # Clean up temp files safely even if an exception occurs
        for path in [temp_image_path, local_excel_path, updated_excel_path]:
            try:
                if os.path.exists(path):
                    os.remove(path)
                    logger.debug(f"Temp file removed: {path}")
            except Exception as e:
                logger.warning(f"Failed to remove temp file {path}: {e}")

# ============================================================
#  Email Utility
# ============================================================
def send_email(recipients, subject, body, attachments=None):
    # attachments: list of dicts, each dict has keys: 'filename' and 'url'

    try:
        # Ensure recipients is always a list
        logger.info(f"Preparing to send email to: {recipients}")
        if isinstance(recipients, str):
            recipients = [recipients]

        msg = Message(subject, recipients=recipients, body=body)

        # Attach files if any
        if attachments:
            for att in attachments:
                url = att.get('url')
                filename = att.get('filename', 'attachment.pdf')
                try:
                    # Normalize Google Drive link to direct download
                    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url or "")
                    if m:
                        file_id = m.group(1)
                        url = f"https://drive.google.com/uc?export=download&id={file_id}"

                    logger.info(f"Fetching attachment: {filename} from {url}")
                    resp = requests.get(url, allow_redirects=True, timeout=15)
                    resp.raise_for_status()
                    msg.attach(filename, "application/pdf", resp.content)
                except Exception as e:
                    logger.error(f"Failed to attach {filename} from {url}: {e}")

        mail.send(msg)
        logger.info("Email sent successfully.")
        return True

    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False

# ============================================================
#  Approval Status Helpers
# ============================================================
def is_already_voided(approval):
    if "Voided" in approval.status:
        return render_template_string(f"""
            <h2 style="text-align: center; color: red;">This request has already been voided.</h2>
            <p style="text-align: center;">Status: {approval.status}</p>
        """)
    return None

def is_already_reviewed(approval, expected_statuses):
    if approval.status not in expected_statuses:
        return render_template_string(f"""
            <h2 style="text-align: center; color: red;">This request has already been reviewed.</h2>
            <p style="text-align: center;">Current Status: {approval.status}</p>
        """)
    return None

# ============================================================
#  API Routes
# ============================================================
@app.route('/get_record/<table>/<id>')
@handle_db_connection
def get_record(table, id):   
    try:
        # Map table names to models
        table_models = {
            'subjects': Subject,
            'rates': Rate,
            'departments': Department,
            'lecturers': Lecturer,
            'programOfficers': ProgramOfficer,
            'heads': Head,
            'others': Other
        }
        
        # Get the appropriate model
        model = table_models.get(table)
        if not model:
            return jsonify({
                'success': False,
                'message': f'Invalid table: {table}'
            }), 400
            
        # Query the record
        record = model.query.get(id)
        if not record:
            return jsonify({
                'success': False,
                'message': f'Record not found in {table} with id {id}'
            }), 404
            
        # Convert record to dictionary
        record_dict = {}
        for column in model.__table__.columns:
            value = getattr(record, column.name)

            # If the table is 'lecturers', use the get_ic_no() to decrypt IC number
            if table == 'lecturers' and column.name == 'ic_no' and value:
                value = record.get_ic_no()
            
            # Convert any non-serializable types to string
            if not isinstance(value, (str, int, float, bool, type(None))):
                value = str(value)
            record_dict[column.name] = value
            
        return jsonify({
            'success': True,
            'record': record_dict
        })
        
    except Exception as e:
        logger.error(f"Error while getting record: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Server error: {str(e)}'
        }), 500

@app.route('/api/check_record_exists/<table>', methods=['GET'])
@handle_db_connection
def check_record_exists(table):
    field = request.args.get('field')
    value = request.args.get('value')

    # Map tables to models and allowed fields (whitelist to avoid injection)
    TABLES = {
        'subjects': (Subject, {'subject_code'}),
        'departments': (Department, {'department_code', 'dean_email'}),
        'lecturers': (Lecturer, {'ic_no', 'email'}),
        'heads': (Head, {'email'}),
        'programOfficers': (ProgramOfficer, {'email'}),
        'others': (Other, {'email'}),
    }

    if table not in TABLES:
        return jsonify({'error': 'Unknown table'}), 400

    model, allowed_fields = TABLES[table]

    if not field or field not in allowed_fields:
        return jsonify({'error': 'Invalid or missing field'}), 400
    if value is None:
        return jsonify({'error': 'Missing value'}), 400

    try:
        exists = model.query.filter(getattr(model, field) == value).first() is not None
        return jsonify({'exists': exists})
    except Exception as e:
        logger.error(f"Error while checking record: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/create_record/<table_type>', methods=['POST'])
@handle_db_connection
def create_record(table_type):
    try:
        data = request.form.to_dict()  

        # ======= Check for Existing Records ========
        if table_type == 'subjects':
            if Subject.query.filter_by(subject_code=data['subject_code']).first():
                return jsonify({'success': False, 'error': f"Subject with code '{data['subject_code']}' already exists."}), 400
    
        elif table_type == 'rates':
            if Rate.query.filter_by(amount=data['amount']).first():
                return jsonify({'success': False, 'error': f"Rate with amount '{data['amount']}' already exists."}), 400
            
        elif table_type == 'departments':
            if Department.query.filter_by(department_code=data['department_code']).first():
                return jsonify({'success': False, 'error': f"Department with code '{data['department_code']}' already exists."}), 400
            
            if Department.query.filter_by(dean_email=data['dean_email']).first():
                return jsonify({'success': False, 'error': f"Department with dean '{data['dean_email']}' already exists."}), 400
                
        elif table_type == 'lecturers':
            # Check if any lecturer already has the same IC number after decrypting it
            ic_no = data['ic_no']  # Get the IC number from the request
            
            lecturer = Lecturer.query.all()  # Retrieve all lecturers from the database
            for existing_lecturer in lecturer:
                decrypted_ic = existing_lecturer.get_ic_no()  # Decrypt the stored IC number
                if decrypted_ic == ic_no:
                    return jsonify({'success': False, 'error': f"Lecturer with IC number '{ic_no}' already exists."}), 400

            # Check if lecturer exists with the same email
            if Lecturer.query.filter_by(email=data['email']).first():
                return jsonify({'success': False, 'error': f"Lecturer with email '{data['email']}' already exists."}), 400

        elif table_type == 'heads':
            if Head.query.filter_by(email=data['email']).first():
                return jsonify({'success': False, 'error': f"Head with email '{data['email']}' already exists."}), 400
            
        elif table_type == 'programOfficers':
            if ProgramOfficer.query.filter_by(email=data['email']).first():
                return jsonify({'success': False, 'error': f"Program Officer with email '{data['email']}' already exists."}), 400  
            
        elif table_type == 'others':
            if Other.query.filter_by(email=data['email']).first():
                return jsonify({'success': False, 'error': f"Entry with email '{data['email']}' already exists."}), 400  
            
        # ======= Record Creation ========
        if table_type == 'subjects':
            new_record = Subject(
                subject_code=data['subject_code'],
                subject_title=data['subject_title'],
                subject_level=data['subject_level'],
                lecture_hours=int(data['lecture_hours']),
                tutorial_hours=int(data['tutorial_hours']),
                practical_hours=int(data['practical_hours']),
                blended_hours=int(data['blended_hours']),
                lecture_weeks=int(data['lecture_weeks']),
                tutorial_weeks=int(data['tutorial_weeks']),
                practical_weeks=int(data['practical_weeks']),
                blended_weeks=int(data['blended_weeks']),
                head_id=data['head_id'],
            )

        elif table_type == 'rates':
            new_record = Rate(
                amount=int(data['amount']),
                status=(None if (s := data.get('status')) in (None, '') else bool(int(s)))
            )
       
        elif table_type == 'departments':
            new_record = Department(
                department_code=data['department_code'],
                department_name=data['department_name'],
                dean_name=data['dean_name'],
                dean_email=data['dean_email']
            )

        elif table_type == 'lecturers':
            new_record = Lecturer(
                name=data['name'],
                email=data['email'],
                password=bcrypt.generate_password_hash('password123$').decode('utf-8'),
                level=data['level'],
                department_id=data['department_id']
            )
            new_record.set_ic_no(data['ic_no'])

        elif table_type == 'heads':
            new_record = Head(
                name=data['name'],
                email=data['email'],
                level=data['level'],
                department_id=data['department_id']
            )

        elif table_type == 'programOfficers':
            new_record = ProgramOfficer(
                name=data['name'],
                email=data['email'],
                password = bcrypt.generate_password_hash('password123$').decode('utf-8'),
                department_id=data['department_id']
            )

        elif table_type == 'others':
            new_record = Other(
                name=data['name'],
                email=data['email'],
                role=data['role']
            )

        else:
            return jsonify({'success': False, 'error': 'Invalid table type'}), 400

        db.session.add(new_record)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'New {table_type[:-1]} created successfully.'
        })

    except IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig)

        if "foreign key constraint fails" in error_msg:
            return jsonify({
                'success': False,
                'error': "One of the linked values is invalid. Please make sure related data (like Department or Head) exists before creating this record."
            }), 400

        elif "Duplicate entry" in error_msg:
            return jsonify({
                'success': False,
                'error': "This record already exists. Please check for duplicate entries."
            }), 400

        return jsonify({
            'success': False,
            'error': "A database integrity error occurred. Please check your input and try again."
        }), 400

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error while creating record: {str(e)}")
        return jsonify({
            'success': False,
            'error': "An unexpected error occurred while creating the record. Please try again."
        }), 500

@app.route('/api/update_record/<table_type>/<id>', methods=['PUT'])
@handle_db_connection
def update_record(table_type, id):
    model_map = {
        'subjects': Subject,
        'departments': Department,
        'lecturers': Lecturer,
        'heads': Head,
        'programOfficers': ProgramOfficer,
        'others': Other
    }

    model = model_map.get(table_type)
    if not model:
        return jsonify({'error': 'Invalid table type'}), 400
    
    try:
        record = model.query.get(id)
        if not record:
            return jsonify({'error': 'Record not found'}), 404
        
        data = request.form.to_dict()  

        # Apply updates for other fields
        for key, value in data.items():
            if hasattr(record, key):
                setattr(record, key, value)

        # Encrypt IC Number before updating
        if 'ic_no' in data:
            record.set_ic_no(data['ic_no'])  # Encrypt the IC number before saving

        db.session.commit()
        return jsonify({'success': True, 'message': 'Record updated successfully.'})
    
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error while updating record: {str(e)}")
        return jsonify({'error': str(e)}), 500
        
@app.route('/api/delete_record/<table_type>', methods=['POST'])
@handle_db_connection
def delete_record(table_type):
    data = request.get_json()
    ids = data.get('ids', [])

    try:
        if table_type == 'subjects':
            Subject.query.filter(Subject.subject_id.in_(ids)).delete()

        elif table_type == 'departments':
            Department.query.filter(Department.department_id.in_(ids)).delete()
    
        elif table_type == 'lecturers':
            Lecturer.query.filter(Lecturer.lecturer_id.in_(ids)).delete()

        elif table_type == 'heads':
            Head.query.filter(Head.head_id.in_(ids)).delete()

        elif table_type == 'programOfficers':
            ProgramOfficer.query.filter(ProgramOfficer.po_id.in_(ids)).delete()
        
        elif table_type == 'others':
            Other.query.filter(Other.other_id.in_(ids)).delete()

        db.session.commit()
        return jsonify({'message': 'Record(s) deleted successfully.'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error while deleting record(s): {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_departments')
@handle_db_connection
def get_departments():
    try:
        departments = Department.query.all()
        return jsonify({
            'success': True,
            'departments': [{'department_id': d.department_id, 
                            'department_code': d.department_code, 
                           'department_name': d.department_name} 
                          for d in departments]
        })
    except Exception as e:
        logger.error(f"Error retrieving departments: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/get_heads')
@handle_db_connection
def get_heads():
    try:
        heads = Head.query.all()
        return jsonify({
            'success': True,
            'heads': [{'head_id': h.head_id,
                             'name': h.name} 
                          for h in heads]
        })
    except Exception as e:
        logger.error(f"Error retrieving heads: {e}")
        return jsonify({'success': False, 'message': str(e)})
    
def delete_requisition_and_attachment(approval_id, suffix):
    # Fetch the approval record first
    approval = RequisitionApproval.query.get(approval_id)
    if not approval:
        logger.warning(f"No approval record found for ID {approval_id}")
        return

    # Rename file
    if approval.file_name:
        name, ext = os.path.splitext(approval.file_name)
        new_file_name = f"{name}_{suffix}{ext}"

        # Update Google Drive file name
        if approval.file_id:
            try:
                drive_service = get_drive_service()
                file_metadata = {"name": new_file_name}
                drive_service.files().update(fileId=approval.file_id, body=file_metadata).execute()
                logger.info(f"Renamed Google Drive file {approval.file_name} -> {new_file_name}")
            except Exception as e:
                logger.error(f"Failed to rename Google Drive file '{approval.file_name}': {e}")

        # Update DB field
        approval.file_name = new_file_name

    # Delete linked LecturerSubject entries
    LecturerSubject.query.filter_by(requisition_id=approval_id).delete(synchronize_session=False)

    # Delete related attachments
    try:
        drive_service = get_drive_service()
        attachments_to_delete = RequisitionAttachment.query.filter_by(requisition_id=approval_id).all()
        for attachment in attachments_to_delete:
            try:
                # Extract file ID from Google Drive URL
                match = re.search(r'/d/([a-zA-Z0-9_-]+)', attachment.attachment_url)
                if not match:
                    logger.warning(f"Invalid Google Drive URL format for attachment {attachment.attachment_name}")
                    continue
                drive_attachment_id = match.group(1)

                # Delete file from Google Drive
                drive_service.files().delete(fileId=drive_attachment_id).execute()

                # Delete attachment record from database
                db.session.delete(attachment)

            except Exception as e:
                logger.error(f"Failed to delete Drive attachment '{attachment.attachment_name}': {e}")
    except Exception as e:
        logger.error(f"Failed to initialize Drive service or delete attachments: {e}")

    # Commit DB changes
    db.session.commit()
