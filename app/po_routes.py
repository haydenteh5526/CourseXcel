import base64, io, logging, os, pytz, re, requests, tempfile
from app import app, db, mail
from app.database import handle_db_connection
from app.models import Admin, ClaimApproval, ClaimAttachment, Department, Head, Lecturer, LecturerClaim, LecturerSubject, Other, ProgramOfficer, Rate, RequisitionApproval, RequisitionAttachment, Subject
from app.excel_generator import generate_requisition_excel
from datetime import datetime
from flask import abort, jsonify, redirect, render_template, render_template_string, request, session, url_for
from flask_mail import Message
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
from sqlalchemy import desc, func

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.route('/poHomepage', methods=['GET', 'POST'])
@handle_db_connection
def poHomepage():
    if 'po_id' not in session:
        return redirect(url_for('loginPage'))

    return render_template('poHomepage.html')

@app.route('/poFormPage', methods=['GET', 'POST'])
@handle_db_connection
def poFormPage():
    if 'po_id' not in session:
        return redirect(url_for('loginPage'))
    
    try:
        # Clean up temp folder first
        cleanup_temp_folder()
        
        # Get all departments and lecturers with their details
        departments = Department.query.all()
        lecturers = Lecturer.query.all()
        
        return render_template('poFormPage.html', 
                             departments=departments,
                             lecturers=lecturers)
    except Exception as e:
        print(f"Error in main route: {str(e)}")
        return str(e), 500
    
@app.route('/get_lecturer_details/<int:lecturer_id>')
@handle_db_connection
def get_lecturer_details(lecturer_id):
    try:
        print(f"Fetching details for lecturer ID: {lecturer_id}")
        lecturer = Lecturer.query.get(lecturer_id)
        
        if not lecturer:
            print(f"Lecturer not found with ID: {lecturer_id}")
            return jsonify({
                'success': False,
                'message': 'Lecturer not found.'
            })
        
        response_data = {
            'success': True,
            'lecturer': {
                'name': lecturer.name,
                'level': lecturer.level,
                'ic_no': lecturer.get_ic_no()
            }
        }
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error getting lecturer details: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        })
    
@app.route('/get_rate_amounts')
@handle_db_connection
def get_rate_amounts():
    try:
        rates = Rate.query.filter(Rate.status.is_(True)).order_by(Rate.amount.asc()).all()
        return jsonify({
            'success': True,
            'rates': [{'rate_id': r.rate_id, 
                       'amount': r.amount} 
                       for r in rates]
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

from sqlalchemy import func

@app.route('/get_assigned_subject/<lecturer_id>')
@handle_db_connection
def get_assigned_subject(lecturer_id):
    try:
        # Get subject_code and the LATEST end date per subject for this lecturer
        rows = (
            db.session.query(
                Subject.subject_code.label('subject_code'),
                func.max(LecturerSubject.end_date).label('end_date')
            )
            .join(LecturerSubject, LecturerSubject.subject_id == Subject.subject_id)
            .filter(LecturerSubject.lecturer_id == lecturer_id)
            .group_by(Subject.subject_code)
            .all()
        )

        # Always return success=True with an array of {subject_code, end_date}
        assigned = []
        for r in rows:
            # Ensure ISO string (or None) for the date
            end_iso = r.end_date.isoformat() if r.end_date else None
            assigned.append({
                'subject_code': r.subject_code,
                'end_date': end_iso
            })

        return jsonify({'success': True, 'assigned': assigned})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e), 'assigned': []})
    
@app.route('/poConversionResult', methods=['POST'])
@handle_db_connection
def poConversionResult():
    if 'po_id' not in session:
        return jsonify(success=False, error="Session expired. Please log in again."), 401
    
    try:
        print("Form Data:", request.form)
        
        # Get basic form data
        department_code = request.form.get('department_code')
        lecturer_id = request.form.get('lecturer_id') 
        name = request.form.get('name')
        designation = request.form.get('designation')
        ic_number = request.form.get('ic_number')
        
        # Helper function to safely convert to int
        def safe_int(value, default=0):
            try:
                if isinstance(value, str):
                    return int(value.strip()) if value.strip() else default
                return int(value)
            except (ValueError, TypeError):
                return default

        # Extract subject-levels early for approval record
        course_details = []
        i = 1
        while True:
            subject_code = request.form.get(f'subjectCode{i}')
            if not subject_code:
                break            

            subject_data = {
                'subject_code': subject_code,
                'subject_title': request.form.get(f'subjectTitle{i}'),
                'subject_level': request.form.get(f'subjectLevel{i}'),
                'start_date': datetime.strptime(request.form.get(f'startDate{i}'), "%Y-%m-%d").date(),
                'end_date': datetime.strptime(request.form.get(f'endDate{i}'), "%Y-%m-%d").date(),
                'lecture_hours': safe_int(request.form.get(f'lectureHours{i}'), 0),
                'lecture_weeks': safe_int(request.form.get(f'lectureWeeks{i}'), 0),
                'tutorial_hours': safe_int(request.form.get(f'tutorialHours{i}'), 0),
                'tutorial_weeks': safe_int(request.form.get(f'tutorialWeeks{i}'), 0),
                'practical_hours': safe_int(request.form.get(f'practicalHours{i}'), 0),
                'practical_weeks': safe_int(request.form.get(f'practicalWeeks{i}'), 0),
                'blended_hours': safe_int(request.form.get(f'blendedHours{i}'), 0),
                'blended_weeks': safe_int(request.form.get(f'blendedWeeks{i}'), 0),
                'hourly_rate': safe_int(request.form.get(f'hourlyRate{i}'), 0)
            }
            course_details.append(subject_data)
            i += 1

        if not course_details:
            return jsonify(success=False, error="No course details provided"), 400

        # Lookup required people for Excel
        po = ProgramOfficer.query.get(session.get('po_id'))
        subject = Subject.query.filter_by(subject_code=request.form.get('subjectCode1')).first()
        head = Head.query.filter_by(head_id=subject.head_id).first()
        department = Department.query.filter_by(department_code=department_code).first()
        ad = Other.query.filter_by(role="Academic Director").first()
        hr = Other.query.filter_by(role="Human Resources").filter(Other.email != "tingting.eng@newinti.edu.my").first()

        # Validate required roles exist
        missing_roles = []

        if not po:
            missing_roles.append("Program Officer")
        if not head:
            missing_roles.append("Head of Programme")
        if not department or not department.dean_name:
            missing_roles.append("Dean")
        if not ad:
            missing_roles.append("Academic Director")
        if not hr:
            missing_roles.append("Human Resources")

        if missing_roles:
            return jsonify(success=False, error=f"Missing required role(s): {', '.join(missing_roles)}"), 400

        po_name = po.name if po else 'N/A'
        head_name = head.name if head else 'N/A'
        dean_name = department.dean_name if department else 'N/A'
        ad_name = ad.name if ad else 'N/A'
        hr_name = hr.name if hr else 'N/A'

        # Generate Excel
        output_path, sign_col = generate_requisition_excel(
            department_code=department_code,
            name=name,
            designation=designation,
            ic_number=ic_number,
            subject_level=request.form.get('subjectLevel1'),
            course_details=course_details,
            po_name=po_name,
            head_name=head_name,
            dean_name=dean_name,
            ad_name=ad_name,
            hr_name=hr_name
        )

        file_name = os.path.basename(output_path)
        file_url, file_id = upload_to_drive(output_path, file_name)

        # Create Approval Record
        approval = RequisitionApproval(
            department_id=department.department_id if department else None,
            lecturer_id=lecturer_id,
            po_id=session.get('po_id'),
            head_id=head.head_id if head else None,
            subject_level=subject.subject_level,
            sign_col=sign_col,
            file_id=file_id,
            file_name=file_name,
            file_url=file_url,
            status="Pending Acknowledgement by PO",
            last_updated=get_current_datetime()
        )
        db.session.add(approval)
        db.session.flush()  # Get approval_id before committing

        approval_id = approval.approval_id

        # Add lecturer_subject entries with requisition_id
        for subject_data in course_details:
            subject_code = subject_data['subject_code']
            subject = Subject.query.filter_by(subject_code=subject_code).first()

            total_lecture_hours = safe_int(subject_data['lecture_hours']) * safe_int(subject_data['lecture_weeks'])
            total_tutorial_hours = safe_int(subject_data['tutorial_hours']) * safe_int(subject_data['tutorial_weeks'])
            total_practical_hours = safe_int(subject_data['practical_hours']) * safe_int(subject_data['practical_weeks'])
            total_blended_hours = safe_int(subject_data['blended_hours']) * safe_int(subject_data['blended_weeks'])

            # Get amount from form and lookup rate_id
            hourly_rate_amount = safe_int(subject_data.get('hourly_rate'), 0)
            rate = Rate.query.filter_by(amount=hourly_rate_amount).first()
            rate_id = rate.rate_id if rate else None

            total_cost = hourly_rate_amount * (
                total_lecture_hours +
                total_tutorial_hours +
                total_practical_hours +
                total_blended_hours
            )

            lecturer_subject = LecturerSubject(
                lecturer_id=lecturer_id,
                requisition_id=approval_id,
                subject_id=subject.subject_id if subject else None,
                start_date=subject_data['start_date'],
                end_date=subject_data['end_date'],
                total_lecture_hours=total_lecture_hours,
                total_tutorial_hours=total_tutorial_hours,
                total_practical_hours=total_practical_hours,
                total_blended_hours=total_blended_hours,
                rate_id=rate_id,
                total_cost=total_cost
            )
            db.session.add(lecturer_subject)

        db.session.commit()

        # ======= Handle Attachments ========
        attachments = request.files.getlist('upload_requisition_attachment')
        attachment_urls = []

        if attachments:
            drive_service = get_drive_service()
            
            for attachment in attachments:
                with tempfile.NamedTemporaryFile(delete=False) as tmp:
                    attachment.save(tmp.name)

                    file_metadata = {'name': attachment.filename}
                    media = MediaFileUpload(tmp.name, mimetype=attachment.mimetype, resumable=True)
                    uploaded = drive_service.files().create(
                        body=file_metadata,
                        media_body=media,
                        fields='id'
                    ).execute()

                    # Set public view permission
                    drive_service.permissions().create(
                        fileId=uploaded['id'],
                        body={'type': 'anyone', 'role': 'reader'},
                    ).execute()

                    file_url = f"https://drive.google.com/file/d/{uploaded['id']}/view"
                    attachment_urls.append((attachment.filename, file_url))
                    os.unlink(tmp.name)

            # Save to ClaimAttachment table
            for filename, url in attachment_urls:
                requisition_attachment = RequisitionAttachment(
                    attachment_name=filename,
                    attachment_url=url,
                    lecturer_id=lecturer_id,
                    requisition_id=approval_id
                )
                db.session.add(requisition_attachment)

            db.session.commit()

        return jsonify(success=True, file_url=file_url)
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in result route: {e}")
        return jsonify(success=False, error=str(e)), 500

@app.route('/poConversionResultPage')
def poConversionResultPage():
    if 'po_id' not in session:
        return redirect(url_for('loginPage'))
        
    approval = RequisitionApproval.query.filter_by(po_id=session.get('po_id')).order_by(RequisitionApproval.approval_id.desc()).first()
    return render_template('poConversionResultPage.html', file_url=approval.file_url)

@app.route('/poRecordsPage', methods=['GET', 'POST'])
@handle_db_connection
def poRecordsPage():
    if 'po_id' not in session:
        return redirect(url_for('loginPage'))

    # Set default tab if none exists
    if 'poRecordsPage_currentTab' not in session:
        session['poRecordsPage_currentTab'] = 'subjects'
    
    subjects = Subject.query.order_by(Subject.subject_code.asc()).all()  
    lecturers = Lecturer.query.order_by(Lecturer.name.asc()).all()
    
    return render_template('poRecordsPage.html', 
                           subjects=subjects,
                           lecturers=lecturers)

@app.route('/set_poRecordsPage_tab', methods=['POST'])
def set_poRecordsPage_tab():
    if 'po_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    session['poRecordsPage_currentTab'] = data.get('poRecordsPage_currentTab')
    return jsonify({'success': True})

@app.route('/poApprovalsPage')
@handle_db_connection
def poApprovalsPage():
    if 'po_id' not in session:
        return redirect(url_for('loginPage'))
    
    # Set default tab if none exists
    if 'poApprovalsPage_currentTab' not in session:
        session['poApprovalsPage_currentTab'] = 'requisitionApprovals'

    po_id = session.get('po_id')

    # lecturers linked to this PO via requisition approvals
    lecturers = (
        db.session.query(Lecturer)
        .join(RequisitionApproval, RequisitionApproval.lecturer_id == Lecturer.lecturer_id)
        .filter(RequisitionApproval.po_id == po_id)
        .order_by(Lecturer.name)
        .distinct()
        .all()
    )
    
    requisitionApprovals = RequisitionApproval.query.filter_by(po_id=po_id)\
                                    .order_by(RequisitionApproval.approval_id.desc())\
                                    .all()
    # Requisition attachments that belong to approvals under this PO
    requisitionAttachments = (
        db.session.query(RequisitionAttachment)
        .join(RequisitionApproval, RequisitionAttachment.requisition_id == RequisitionApproval.approval_id)
        .filter(RequisitionApproval.po_id == po_id)
        .order_by(RequisitionAttachment.attachment_id.desc())
        .all()
    )

    claimApprovals = ClaimApproval.query.filter_by(po_id=po_id)\
                                    .order_by(ClaimApproval.approval_id.desc())\
                                    .all()
    # Claim attachments that belong to approvals under this PO
    claimAttachments = (
        db.session.query(ClaimAttachment)
        .join(ClaimApproval, ClaimAttachment.claim_id == ClaimApproval.approval_id)
        .filter(ClaimApproval.po_id == po_id)
        .order_by(ClaimAttachment.attachment_id.desc())
        .all()
    )
      
    # Get all LecturerSubject records linked to completed requisitions
    subjects = (
        db.session.query(
            LecturerSubject,
            Lecturer,
            Subject.subject_code,
            Subject.subject_title,
            Subject.subject_level
        )
        .join(Lecturer, LecturerSubject.lecturer_id == Lecturer.lecturer_id)
        .join(Subject, LecturerSubject.subject_id == Subject.subject_id)
        .join(RequisitionApproval, LecturerSubject.requisition_id == RequisitionApproval.approval_id)
        .filter(
            RequisitionApproval.status == 'Completed',
            RequisitionApproval.po_id == po_id
        )
        .order_by(desc(RequisitionApproval.approval_id))
        .all()
    )

    claimDetails = []
    for ls, lecturer, code, title, level in subjects:
        # Sum all claimed hours for this lecturer/subject
        claimed = (
            db.session.query(
                func.coalesce(func.sum(LecturerClaim.lecture_hours), 0),
                func.coalesce(func.sum(LecturerClaim.tutorial_hours), 0),
                func.coalesce(func.sum(LecturerClaim.practical_hours), 0),
                func.coalesce(func.sum(LecturerClaim.blended_hours), 0)
            )
            .filter_by(lecturer_id=lecturer.lecturer_id, subject_id=ls.subject_id)
            .first()
        )

        remaining = {
            'lecturer': lecturer,
            'subject_code': code,
            'subject_title': title,
            'subject_level': level,
            'start_date': ls.start_date,
            'end_date': ls.end_date,
            'hourly_rate': ls.rate.amount if ls.rate else 0,
            'lecture_hours': ls.total_lecture_hours - claimed[0],
            'tutorial_hours': ls.total_tutorial_hours - claimed[1],
            'practical_hours': ls.total_practical_hours - claimed[2],
            'blended_hours': ls.total_blended_hours - claimed[3],
        }
        claimDetails.append(remaining)

    return render_template('poApprovalsPage.html', 
                           lecturers=lecturers,
                           requisitionApprovals=requisitionApprovals,
                           requisitionAttachments=requisitionAttachments,
                           claimApprovals=claimApprovals,
                           claimAttachments=claimAttachments,
                           claimDetails=claimDetails)

@app.route('/set_poApprovalsPage_tab', methods=['POST'])
def set_poApprovalsPage_tab():
    if 'po_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    session['poApprovalsPage_currentTab'] = data.get('poApprovalsPage_currentTab')
    return jsonify({'success': True})

@app.route('/check_requisition_status/<int:approval_id>')
def check_requisition_status(approval_id):
    approval = RequisitionApproval.query.get_or_404(approval_id)
    return jsonify({'status': approval.status})

@app.route('/api/po_review_requisition/<approval_id>', methods=['POST'])
@handle_db_connection
def po_review_requisition(approval_id):
    try:
        data = request.get_json()
        signature_data = data.get("image")

        if not signature_data or "," not in signature_data:
            return jsonify(success=False, error="Invalid image data format")

        # Fetch approval record
        approval = RequisitionApproval.query.get(approval_id)
        if not approval:
            return jsonify(success=False, error="Approval record not found")

        # Process signature and upload updated file
        process_signature_and_upload(approval, signature_data, "B")

        # Update status after signature inserted
        approval.status = "Pending Acknowledgement by HOP"
        approval.last_updated = get_current_datetime()
        db.session.commit()

        try:
            notify_approval(approval, approval.head.email if approval.head else None, "head_review_requisition", "Head of Programme")
        except Exception as e:
            logging.error(f"Failed to notify HOP: {e}")    

        return jsonify(success=True)

    except Exception as e:
        logging.error(f"Error uploading signature: {e}")
        return jsonify(success=False, error=str(e)), 500
    
@app.route('/api/head_review_requisition/<approval_id>', methods=['GET', 'POST'])
def head_review_requisition(approval_id):
    approval = RequisitionApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    # Check if already voided
    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    # Check if already reviewed
    reviewed_response = is_already_reviewed(approval, ["Pending Acknowledgement by HOP"])
    if reviewed_response:
        return reviewed_response

    # Otherwise render the review page
    if request.method == 'GET':
        return render_template("reviewRequisitionApprovalRequest.html", approval=approval)

    # POST logic
    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "E")
            approval.status = "Pending Acknowledgement by Dean / HOS"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            try:
                notify_approval(approval, approval.department.dean_email if approval.department else None, "dean_review_requisition", "Dean / HOS")
            except Exception as e:
                logging.error(f"Failed to notify Dean: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by HOP - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_requisition_and_attachment(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("HOP", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")

        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400
  
@app.route('/api/dean_review_requisition/<approval_id>', methods=['GET', 'POST'])
def dean_review_requisition(approval_id):
    approval = RequisitionApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    # Check if already voided
    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    # Check if already reviewed
    reviewed_response = is_already_reviewed(approval, ["Pending Acknowledgement by Dean / HOS"])
    if reviewed_response:
        return reviewed_response

    # Otherwise render the review page
    if request.method == 'GET':
        return render_template("reviewRequisitionApprovalRequest.html", approval=approval)
    
    # POST logic
    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "G")
            approval.status = "Pending Acknowledgement by Academic Director"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            ad = Other.query.filter_by(role="Academic Director").first()

            try:
                notify_approval(approval, ad.email if ad else None, "ad_review_requisition", "Academic Director")
            except Exception as e:
                logging.error(f"Failed to notify AD: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by Dean / HOS - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_requisition_and_attachment(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("Dean", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")

        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400
        
@app.route('/api/ad_review_requisition/<approval_id>', methods=['GET', 'POST'])
def ad_review_requisition(approval_id):
    approval = RequisitionApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    # Check if already voided
    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    # Check if already reviewed
    reviewed_response = is_already_reviewed(approval, ["Pending Acknowledgement by Academic Director"])
    if reviewed_response:
        return reviewed_response

    # Otherwise render the review page
    if request.method == 'GET':
        return render_template("reviewRequisitionApprovalRequest.html", approval=approval)
    
    action = request.form.get('action')
    if action == 'approve':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "I")
            approval.status = "Pending Acknowledgement by HR"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            hr = Other.query.filter(Other.role == "Human Resources", Other.email != "tingting.eng@newinti.edu.my").first()

            try:
                notify_approval(approval, hr.email if hr else None, "hr_review_requisition", "HR")
            except Exception as e:
                logging.error(f"Failed to notify HR: {e}")

            return '''<script>alert("Request approved successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500

    elif action == 'reject':
        reason = request.form.get('reject_reason')
        if not reason:
            return "Rejection reason required", 400
        
        approval.status = f"Rejected by Academic Director - {reason.strip()}"
        approval.last_updated = get_current_datetime()

        # Delete related records and rename approval file
        delete_requisition_and_attachment(approval.approval_id, "REJECTED")

        try:
            send_rejection_email("AD", approval, reason.strip())
        except Exception as e:
            logging.error(f"Failed to send rejection email: {e}")
        
        return '''<script>alert("Request rejected successfully. You may now close this window.")</script>'''
    
    return "Invalid action", 400
        
@app.route('/api/hr_review_requisition/<approval_id>', methods=['GET', 'POST'])
def hr_review_requisition(approval_id):
    approval = RequisitionApproval.query.get(approval_id)
    if not approval:
        abort(404, description="Approval record not found")

    # Check if already voided
    voided_response = is_already_voided(approval)
    if voided_response:
        return voided_response

    # Check if already reviewed
    reviewed_response = is_already_reviewed(approval, ["Pending Acknowledgement by HR"])
    if reviewed_response:
        return reviewed_response

    # Otherwise render the review page
    if request.method == 'GET':
        return render_template("receiveRequisitionApprovalRequest.html", approval=approval)

    action = request.form.get('action')
    if action == 'confirm':
        try:
            process_signature_and_upload(approval, request.form.get('signature_data'), "K")
            approval.status = "Completed"
            approval.last_updated = get_current_datetime()
            db.session.commit()

            try:
                subject = f"Part-time Lecturer Requisition Approval Request Completed - {approval.lecturer.name} ({approval.subject_level})"
                body = (
                    f"Dear All,\n\n"
                    f"The part-time lecturer requisition request has been fully approved by all parties.\n\n"
                    f"Please click the link below to access the final approved file:\n"
                    f"{approval.file_url}\n\n"
                    "Thank you for your cooperation.\n\n"
                    "Best regards,\n"
                    "The CourseXcel Team"
                )
                
                # Get final HR and admin
                final_hr_email = "tingting.eng@newinti.edu.my"
                admin = Admin.query.filter_by(admin_id=1).first()

                # Base recipients from related models
                recipients = [
                    approval.program_officer.email if approval.program_officer else None,
                    approval.head.email if approval.head else None,
                    approval.department.dean_email if approval.department else None,
                ]

                # Get "Other" roles
                ad = Other.query.filter_by(role="Academic Director").first()
                hr = Other.query.filter_by(role="Human Resources").filter(Other.email != final_hr_email).first()

                # Append AD and first HR if exists
                if ad and ad.email:
                    recipients.append(ad.email)
                if hr and hr.email:
                    recipients.append(hr.email)

                # Append final HR and admin
                recipients.append(final_hr_email)
                if admin and admin.email:
                    recipients.append(admin.email)

                # Filter out any Nones or duplicates
                recipients = list(filter(None, set(recipients)))

                send_email(recipients, subject, body)
            except Exception as e:
                logging.error(f"Failed to notify All: {e}")

            return '''<script>alert("Request confirmed successfully. You may now close this window.")</script>'''
        except Exception as e:
            return str(e), 500
    
    return "Invalid action", 400

@app.route('/api/void_requisition/<approval_id>', methods=['POST'])
@handle_db_connection
def void_requisition(approval_id):
    try:
        data = request.get_json()
        reason = data.get("reason", "").strip()

        if not reason:
            return jsonify(success=False, error="Reason for voiding is required."), 400

        approval = RequisitionApproval.query.get(approval_id)
        if not approval:
            return jsonify(success=False, error="Approval record not found."), 404

        current_status = approval.status

        # Allow voiding only at specific stages
        if current_status in [
            "Pending Acknowledgement by PO",
            "Pending Acknowledgement by HOP",
            "Pending Acknowledgement by Dean / HOS",
            "Pending Acknowledgement by Academic Director",
            "Pending Acknowledgement by HR"
        ]:
            approval.status = f"Voided - {reason}"
            approval.last_updated = get_current_datetime()

            # Delete related records and rename approval file
            delete_requisition_and_attachment(approval.approval_id, "VOIDED")

        else:
            return jsonify(success=False, error="Request cannot be voided at this stage."), 400

        ad = Other.query.filter_by(role="Academic Director").first()

        # Determine recipients based on current stage
        recipients = []
        if current_status == "Pending Acknowledgement by Dean / HOS":
            recipients = [approval.head.email]
        elif current_status == "Pending Acknowledgement by Academic Director":
            recipients = [approval.head.email, approval.department.dean_email]
        elif current_status == "Pending Acknowledgement by HR":
            recipients = [approval.head.email, approval.department.dean_email, ad.email]

        recipients = list(set(filter(None, recipients)))  # Remove duplicates and None

        # Send notification emails
        subject = f"Part-time Lecturer Requisition Request Voided - {approval.lecturer.name} ({approval.subject_level})"
        body = (
            f"Dear All,\n\n"
            f"The part-time lecturer requisition request has been voided by the Requester.\n"
            f"Reason: {reason}\n\n"
            f"Please review the file here:\n{approval.file_url}\n"
            "Please do not take any further action on this request.\n\n"
            "Thank you,\n"
            "The CourseXcel Team"
        )

        if recipients:
            success = send_email(recipients, subject, body)
            if not success:
                logging.error(f"Failed to send void notification email to: {recipients}")
        
        return jsonify(success=True)

    except Exception as e:
        logging.error(f"Error voiding requisition: {e}")
        return jsonify(success=False, error="Internal server error."), 500

@app.route('/poProfilePage')
def poProfilePage():
    po_email = session.get('po_email')
    if not po_email:
        return redirect(url_for('loginPage'))

    po = ProgramOfficer.query.filter_by(email=po_email).first()

    return render_template('poProfilePage.html', po=po)

def cleanup_temp_folder():
    """Clean up all files in the temp folder"""
    temp_folder = os.path.join(app.root_path, 'temp')
    if os.path.exists(temp_folder):
        for filename in os.listdir(temp_folder):
            file_path = os.path.join(temp_folder, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    logger.info(f"Cleaned up file: {file_path}")
            except Exception as e:
                logger.error(f"Error cleaning up file {file_path}: {e}")

def delete_file(file_path):
    """Helper function to delete a file"""
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"File deleted successfully: {file_path}")
            return True
    except Exception as e:
        logger.error(f"Error deleting file {file_path}: {e}")
        return False

def get_drive_service():
    SERVICE_ACCOUNT_FILE = '/home/TomazHayden/coursexcel-459515-3d151d92b61f.json'
    SCOPES = ['https://www.googleapis.com/auth/drive']
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

def upload_to_drive(file_path, file_name):
    try:
        service = get_drive_service()

        file_metadata = {
            'name': file_name,
            'mimeType': 'application/vnd.google-apps.spreadsheet'  # Convert to Google Sheets
        }

        media = MediaFileUpload(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )

        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()

        # Make file publicly accessible
        service.permissions().create(
            fileId=file.get('id'),
            body={'type': 'anyone', 'role': 'reader'},
        ).execute()

        file_id = file.get('id')
        file_url = f"https://docs.google.com/spreadsheets/d/{file_id}/edit"

        return file_url, file_id

    except Exception as e:
        logging.error(f"Failed to upload to Google Drive: {e}")
        raise

def download_from_drive(file_id):
    drive_service = get_drive_service()

    request = drive_service.files().export_media(fileId=file_id,
        mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    output_folder = os.path.join(os.path.abspath(os.path.dirname(__file__)), "temp")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    local_path = os.path.join(output_folder, f"{file_id}.xlsx")

    fh = io.FileIO(local_path, 'wb')
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    fh.close()
    return local_path

def get_current_datetime():
    malaysia_tz = pytz.timezone('Asia/Kuala_Lumpur')
    now_myt = datetime.now(malaysia_tz)
    return now_myt.strftime('%a, %d %b %y, %I:%M:%S %p')
          
def save_signature_image(signature_data, approval_id, temp_folder):
    try:
        header, encoded = signature_data.split(",", 1)
        binary_data = base64.b64decode(encoded)
        image = Image.open(BytesIO(binary_data))
        temp_image_path = os.path.join(temp_folder, f"signature_{approval_id}.png")
        image.save(temp_image_path)
        return temp_image_path
    except Exception as e:
        logging.error(f"Signature decoding error: {e}")
        return None

def insert_signature_and_date(local_excel_path, signature_path, cell_prefix, row, updated_path):
    wb = load_workbook(local_excel_path)
    ws = wb['Sheet1']

    # Insert signature
    sign_cell = f"{cell_prefix}{row}"
    signature_img = ExcelImage(signature_path)
    signature_img.width = 100
    signature_img.height = 30
    ws.add_image(signature_img, sign_cell)

    # Insert date
    date_cell = f"{cell_prefix}{row + 3}"
    malaysia_time = datetime.now(pytz.timezone('Asia/Kuala_Lumpur'))
    ws[date_cell] = f"Date: {malaysia_time.strftime('%d/%m/%Y')}"

    wb.save(updated_path)

def process_signature_and_upload(approval, signature_data, col_letter):
    temp_folder = os.path.join("temp")
    os.makedirs(temp_folder, exist_ok=True)

    temp_image_path = save_signature_image(signature_data, approval.approval_id, temp_folder)
    if not temp_image_path:
        raise ValueError("Invalid signature image data")

    local_excel_path = download_from_drive(approval.file_id)
    updated_excel_path = os.path.join(temp_folder, approval.file_name)

    try:
        insert_signature_and_date(local_excel_path, temp_image_path, col_letter, approval.sign_col, updated_excel_path)
        new_file_url, new_file_id = upload_to_drive(updated_excel_path, approval.file_name)

        # Delete old versions from Drive except the new one
        drive_service = get_drive_service()

        # Delete old file by stored file ID (approval.file_id) if different from new_file_id
        if approval.file_id and approval.file_id != new_file_id:
            try:
                drive_service.files().delete(fileId=approval.file_id).execute()
                logging.info(f"Deleted old file with ID {approval.file_id}")
            except Exception as e:
                logging.warning(f"Failed to delete old file {approval.file_id}: {e}")
                
        # Update DB record
        approval.file_url = new_file_url
        approval.file_id = new_file_id
        approval.last_updated = get_current_datetime()
        db.session.commit()

    finally:
        # Clean up temp files safely even if an exception occurs
        for path in [temp_image_path, local_excel_path, updated_excel_path]:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception as e:
                logging.warning(f"Failed to remove temp file {path}: {e}")

def delete_requisition_and_attachment(approval_id, suffix):
    # Fetch the approval record first
    approval = RequisitionApproval.query.get(approval_id)
    if not approval:
        logging.warning(f"No approval record found for ID {approval_id}")
        return

    # Rename file
    if approval.file_name:
        name, ext = os.path.splitext(approval.file_name)
        new_file_name = f"{name}_{suffix}{ext}"

        # Update Google Drive file name
        if approval.file_id:
            try:
                drive_service = get_drive_service()
                file_metadata = {"name": new_file_name}
                drive_service.files().update(fileId=approval.file_id, body=file_metadata).execute()
                logging.info(f"Renamed Google Drive file {approval.file_name} -> {new_file_name}")
            except Exception as e:
                logging.error(f"Failed to rename Google Drive file '{approval.file_name}': {e}")

        # Update DB field
        approval.file_name = new_file_name

    # Delete linked LecturerSubject entries
    LecturerSubject.query.filter_by(requisition_id=approval_id).delete(synchronize_session=False)

    # Delete related attachments
    try:
        drive_service = get_drive_service()
        attachments_to_delete = RequisitionAttachment.query.filter_by(requisition_id=approval_id).all()
        for attachment in attachments_to_delete:
            try:
                # Extract file ID from Google Drive URL
                match = re.search(r'/d/([a-zA-Z0-9_-]+)', attachment.attachment_url)
                if not match:
                    logging.warning(f"Invalid Google Drive URL format for attachment {attachment.attachment_name}")
                    continue
                drive_attachment_id = match.group(1)

                # Delete file from Google Drive
                drive_service.files().delete(fileId=drive_attachment_id).execute()

                # Delete attachment record from database
                db.session.delete(attachment)

            except Exception as e:
                logging.error(f"Failed to delete Drive attachment '{attachment.attachment_name}': {e}")
    except Exception as e:
        logging.error(f"Failed to initialize Drive service or delete attachments: {e}")

    # Commit DB changes
    db.session.commit()

def is_already_voided(approval):
    if "Voided" in approval.status:
        return render_template_string(f"""
            <h2 style="text-align: center; color: red;">This request has already been voided.</h2>
            <p style="text-align: center;">Status: {approval.status}</p>
        """)
    return None

def is_already_reviewed(approval, expected_statuses):
    if any(status in approval.status for status in expected_statuses):
        return render_template_string(f"""
            <h2 style="text-align: center; color: red;">This request has already been reviewed.</h2>
            <p style="text-align: center;">Status: {approval.status}</p>
        """)
    return None

def get_requisition_attachments(approval_id):
    # Returns a list of RequisitionAttachment objects
    return RequisitionAttachment.query.filter_by(requisition_id=approval_id).all()

def send_email(recipients, subject, body, attachments=None):
    try:
        if isinstance(recipients, str):
            recipients = [recipients]

        msg = Message(subject, recipients=recipients, body=body)

        if attachments:
            for att in attachments:
                url = att.get('url')
                filename = att.get('filename', 'attachment.pdf')
                try:
                    # Normalize Google Drive link to direct download
                    m = re.search(r"/d/([a-zA-Z0-9_-]+)", url or "")
                    if m:
                        file_id = m.group(1)
                        url = f"https://drive.google.com/uc?export=download&id={file_id}"

                    app.logger.info(f"Fetching attachment: {filename} from {url}")
                    resp = requests.get(url, allow_redirects=True, timeout=15)
                    resp.raise_for_status()
                    msg.attach(filename, "application/pdf", resp.content)
                except Exception as e:
                    app.logger.error(f"Failed to attach {filename} from {url}: {e}. Skipping this attachment.")
                    # continue without failing the whole email

        app.logger.info("Sending email via SMTP…")
        mail.send(msg)
        app.logger.info("Email sent.")
        return True

    except Exception as e:
        app.logger.error(f"Failed to send email (SMTP or earlier): {e}")
        return False
    
def notify_approval(approval, recipient_email, next_review_route, greeting):
    if not recipient_email:
        logging.error("No recipient email provided for approval notification.")
        return

    review_url = url_for(next_review_route, approval_id=approval.approval_id, _external=True)

    # Get all attachments for this approval
    attachments = get_requisition_attachments(approval.approval_id)

    # Convert attachments to list of dicts with filename & URL
    attachment_list = [
        {'filename': att.attachment_name, 'url': att.attachment_url}
        for att in attachments
    ]

    if greeting == "HR":
        subject = f"Part-time Lecturer Requisition Acknowledgement Required - {approval.lecturer.name} ({approval.subject_level})"
        body = (
            f"Dear {greeting},\n\n"
            f"The part-time lecturer requisition form has been fully approved and is now ready for your acknowledgement.\n\n"
            f"To confirm receipt, kindly click the link below and provide your digital signature:\n"
            f"{review_url}\n\n"
            "Attachments are included for your reference.\n\n"
            "Thank you,\n"
            "The CourseXcel Team"
        )
    else:
        subject = f"Part-time Lecturer Requisition Approval Request - {approval.lecturer.name} ({approval.subject_level})"
        body = (
            f"Dear {greeting},\n\n"
            f"There is a part-time lecturer requisition request pending your review and approval.\n\n"
            f"Please click the link below to review and approve or reject the request:\n"
            f"{review_url}\n\n"
            "Attachments are included for your reference.\n\n"
            "Thank you,\n"
            "The CourseXcel Team"
        )

    send_email(recipient_email, subject, body, attachments=attachment_list)

def send_rejection_email(role, approval, reason):
    subject = f"Part-time Lecturer Requisition Request Rejected - {approval.lecturer.name} ({approval.subject_level})"

    role_names = {
        "HOP": "Head of Programme",
        "Dean": "Dean / Head of School",
        "AD": "Academic Director",
        "HR": "HR"
    }

    ad = Other.query.filter_by(role="Academic Director").first()

    recipients_map = {
        "HOP": [approval.program_officer.email] if approval.program_officer else [],
        "Dean": [
            approval.program_officer.email if approval.program_officer else None,
            approval.head.email if approval.head else None
        ],
        "AD": [
            approval.program_officer.email if approval.program_officer else None,
            approval.head.email if approval.head else None,
            approval.department.dean_email if approval.department else None
        ],
        "HR": [
            approval.program_officer.email if approval.program_officer else None,
            approval.head.email if approval.head else None,
            approval.department.dean_email if approval.department else None,
            ad.email if ad else None
        ]
    }

     # Clean up None values and deduplicate
    recipients = list(set(filter(None, recipients_map.get(role, []))))

    rejected_by = role_names.get(role, "Unknown Role")
    greeting = "Dear Requester" if role == "HOP" else "Dear All"

    body = (
        f"{greeting},\n\n"
        f"The part-time lecturer requisition approval request has been rejected by the {rejected_by}.\n\n"
        f"Reason for rejection: {reason}\n\n"
        f"You can review the file here:\n{approval.file_url}\n\n"
        "Please take necessary actions.\n\n"
        "Thank you,\n"
        "The CourseXcel Team"
    )

    send_email(recipients, subject, body)
